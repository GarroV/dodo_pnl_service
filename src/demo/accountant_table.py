"""
Собрать «таблицу бухгалтера» в формате PLATA из демо-данных.

Экран сверки в демо-стенде должен показывать работающий разбор чужого файла —
а не картинку. Значит демо обязано **само** производить xlsx, который читает
`payroll.importers.plata_xlsx.read_plata_file`, и производить его из тех же
данных, что показаны на экранах продукта (иначе сверка сравнивала бы одно
с другим и врала бы посетителю про «сошлось»).

Формат — источник истины `plata_xlsx.py`: имена листов (`SHEET_MAP`), точные
заголовки колонок (`FIELDS`, `GROSS_HEADER`) и то, что разбор считает числом
(`_numeric`, `NUMBER`). Этот модуль их не копирует руками, а импортирует —
разъедутся форматы, тесты упадут сразу, а не после жалобы на демо.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal

import openpyxl

from payroll.importers.plata_xlsx import FIELDS, GROSS_HEADER, SHEET_MAP

# Порядок колонок листа: сначала обязательные поля расчёта, затем то, что
# бухгалтерия считает сама. exp_gross — последним и с заголовком, который
# зависит от схемы: `_map_columns` в разборе строит колонки ровно в этом же
# составе (FIELDS + отдельно посчитанный exp_gross), порядок здесь — только
# для читаемости файла человеком, на разбор не влияет.
_COLUMN_ORDER: list[tuple[str, str]] = [
    ("coefficient", "coefficient"),
    ("base_rate", "base_rate"),
    ("insured", "insured"),
    ("regular", "regular"),
    ("holiday", "holiday"),
    ("vacation", "vacation"),
    ("sick", "sick"),
    ("correction", "correction"),
    ("deduction", "deduction"),
    ("cash", "cash"),
    ("meal", "meal"),
    ("net", "exp_net"),
    ("contributions", "exp_contrib"),
    ("total_cost", "exp_total"),
    ("gross", "exp_gross"),
]


@dataclass(frozen=True)
class TableRow:
    """Одна строка таблицы бухгалтера: человек, его входы и посчитанные им итоги."""

    first: str
    last: str
    sheet: str                  # точное имя листа из SHEET_MAP, включая пробелы
    scheme: str                 # standard | half_time | half_time_min_base | temporary
    coefficient: Decimal
    base_rate: Decimal
    insured: Decimal
    regular: Decimal
    holiday: Decimal
    vacation: Decimal
    sick: Decimal
    deduction: Decimal
    cash: Decimal
    correction: Decimal | None  # None — колонка пустая
    meal: Decimal | None        # None — колонка пустая
    net: Decimal | None
    gross: Decimal | None
    contributions: Decimal | None
    total_cost: Decimal | None


def _header_for(field_key: str, scheme: str) -> str:
    """Заголовок колонки — тот же, который ищет разбор (первое имя из списка)."""
    names = GROSS_HEADER[scheme] if field_key == "exp_gross" else FIELDS[field_key]
    return names[0]


def _check_scheme(row: TableRow) -> str:
    """Схема листа определяется листом, а не строкой (см. `SHEET_MAP`).

    Если строка приписана к листу с чужой схемой, файл всё равно прочитается —
    просто заголовок бруто выйдет не тот, что ждёт разбор этой схемы, и
    колонка «не найдётся» молча. Дешевле уронить сборку явной ошибкой, чем
    отдать демо файл, который тихо не сходится.
    """
    if row.sheet not in SHEET_MAP:
        raise ValueError(
            f"лист {row.sheet!r} не входит в формат PLATA (см. SHEET_MAP)"
        )
    expected_scheme, _group = SHEET_MAP[row.sheet]
    if row.scheme != expected_scheme:
        raise ValueError(
            f"{row.first} {row.last}: схема {row.scheme!r} не подходит листу "
            f"{row.sheet!r} — на нём считают по схеме {expected_scheme!r}"
        )
    return expected_scheme


def build_accountant_table(rows: list[TableRow]) -> bytes:
    """Собрать книгу xlsx в формате PLATA. Возвращает содержимое файла байтами.

    Числа пишутся как `Decimal` напрямую, а не переводятся во `float` руками:
    openpyxl сам сериализует `Decimal` в XML как десятичный текст (проверено
    прогоном пары сотен тысяч случайных сумм с копейками через реальное
    сохранение/чтение книги — расхождений не нашлось), поэтому копейки не
    едут ни на записи, ни на последующем чтении `read_plata_file`.
    """
    # Строки группируются по листу в порядке первого появления — лист заводится,
    # только если для него есть строки (иначе файл обрастал бы пустыми листами,
    # которых разбор не ждёт).
    by_sheet: dict[str, list[TableRow]] = {}
    for row in rows:
        _check_scheme(row)
        by_sheet.setdefault(row.sheet, []).append(row)

    wb = openpyxl.Workbook()
    for sheet_name, sheet_rows in by_sheet.items():
        scheme = sheet_rows[0].scheme
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(1, 1).value = "RB"
        ws.cell(1, 2).value = "IME"
        ws.cell(1, 3).value = "PREZIME"
        for col_offset, (_attr, field_key) in enumerate(_COLUMN_ORDER):
            ws.cell(1, 4 + col_offset).value = _header_for(field_key, scheme)

        for index, row in enumerate(sheet_rows, start=1):
            ws.cell(index + 1, 1).value = index  # число — иначе строка не пронумерована
            ws.cell(index + 1, 2).value = row.first
            ws.cell(index + 1, 3).value = row.last
            for col_offset, (attr, _field_key) in enumerate(_COLUMN_ORDER):
                ws.cell(index + 1, 4 + col_offset).value = getattr(row, attr)

    # Workbook() создаёт лист «Sheet» по умолчанию — мы его не заполняли,
    # и в готовой книге ему делать нечего.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
