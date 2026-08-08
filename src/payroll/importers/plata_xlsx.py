"""
Импорт зарплатной таблицы партнёра (Сербия, формат PLATA).

Нужен для двух вещей: перенести существующие данные при подключении партнёра
и сверять движок с их расчётом. Формат сербский, поэтому лежит в importers —
у другого партнёра будет свой.

**Разбор обязан быть громким (T021).** Файл партнёра — это чужая таблица,
которую правят руками: лист переименуют, колонку добавят, в часы напишут
словами. Раньше всё это проходило молча: чужой лист пропускался, ненайденная
колонка становилась нулём, а строки ниже 19-й не читались вовсе. Молчание здесь
дороже отказа: неполный табель посчитается и даст правдоподобно неверную
зарплату. Поэтому разбор возвращает не только строки, но и **находки**
(`Finding`) — всё, что он не смог разобрать и потому не загрузил.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import openpyxl

from ..engine import Employee, Timesheet, d

# лист → (схема расчёта, группа)
SHEET_MAP: dict[str, tuple[str, str]] = {
    "NS  kancelarija ":               ("standard",           "office"),
    "BG1 pun obracun ":               ("standard",           "kitchen"),
    "NS 1 Bulevar ":                  ("standard",           "kitchen"),
    "NS 2 Dunavska ":                 ("standard",           "kitchen"),
    "BG1  pola radnog  vremena":      ("half_time",          "kitchen"),
    "BG 1 pola radnog vremena  puno": ("half_time_min_base", "kitchen"),
    "NS pola radnog  vremena puno ":  ("half_time_min_base", "kitchen"),
    "NS privremeni poslovi ":         ("temporary",          "temporary"),
}

# точные заголовки, первый найденный выигрывает
FIELDS: dict[str, list[str]] = {
    "coefficient": ["KOEFICIJENT"],
    "base_rate":   ["CENA RADA"],
    "insured":     ["UKUPNO SATI ZA OBRACUN DOPRINOSA", "UKUPNO SATI"],
    "regular":     ["SATI RADA ZA OBRACUN NETA", "SATI RADA"],
    "holiday":     ["SAT NA RAD PRAZNIKA"],
    "vacation":    ["GODISNJI ODMOR"],
    "sick":        ["SATI BOLOVANJE"],
    "correction":  ["KOREKCIJA DO MINIMALCA"],
    "deduction":   ["OBUSTAVA"],
    "cash":        ["ISPLATA U KES"],
    "meal":        ["TOPLI OBROK I REGRES"],
    "exp_net":     ["UKUPNO ZA ISPLATU"],
    "exp_contrib": ["DOPRINOSI"],
    "exp_total":   ["UKUPAN TROSAK PO RADNIKU"],
}

# «бруто» в разных схемах называется по-разному
GROSS_HEADER: dict[str, list[str]] = {
    "standard":           ["BRUTO"],
    "half_time":          ["BRUTO DOPRINOSI"],
    "half_time_min_base": ["BRUTO ZA POREZ"],
    "temporary":          ["BRUTO"],
}

HOUR_FIELDS = ("regular", "holiday", "vacation", "sick")


# Строка, которую разбор считает числом. Нарочно узко: «8,5» здесь не
# принимается, потому что в чужой таблице запятая с равной вероятностью
# десятичная и разделитель тысяч, и угадывать — значит однажды прочитать
# «1,234» как 1,234 часа вместо 1234.
NUMBER = re.compile(r"^-?\d+(\.\d+)?$")


@dataclass
class ImportedRow:
    """Строка таблицы: входные данные плюс то, что посчитала бухгалтерия."""
    sheet: str
    scheme: str
    group: str
    name: str
    employee: Employee
    timesheet: Timesheet
    sheet_meal: Decimal | None
    expected: dict[str, Decimal | None]


@dataclass(frozen=True)
class Finding:
    """Что разбор не смог прочитать и потому не загрузил.

    `kind` — чтобы отчёт мог сгруппировать однородное, а не вываливать простыню:
    `sheet` (лист), `column` (колонка), `row` (строка), `value` (значение).
    `where` — место в файле теми же словами, какими его видит человек: имя листа
    и номер строки. Без него отчёт сообщает о беде, но не о том, куда идти.
    """
    kind: str
    where: str
    text: str


@dataclass
class PlataFile:
    """Результат разбора: что прочитано и что не прочитано."""
    rows: list[ImportedRow] = field(default_factory=list)
    findings: list[Finding] = field(default_factory=list)


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value)).strip().upper()


def _map_columns(ws, scheme: str) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, 45):
        value = ws.cell(1, col).value
        if value and _norm(value) not in headers:
            headers[_norm(value)] = col

    fields = dict(FIELDS, exp_gross=GROSS_HEADER[scheme])
    mapped: dict[str, int] = {}
    for key, names in fields.items():
        for name in names:
            if name in headers:
                mapped[key] = headers[name]
                break
    return mapped


def _title_of(key: str, scheme: str) -> str:
    """Как колонка называется в самой таблице — по нему её ищут глазами."""
    names = GROSS_HEADER[scheme] if key == "exp_gross" else FIELDS.get(key, [])
    return names[0] if names else key


def _numeric(raw) -> bool:
    """Годится ли значение ячейки в число. Пусто — годится (это ноль)."""
    if raw is None:
        return True
    if isinstance(raw, bool):
        return False
    if isinstance(raw, (int, float, Decimal)):
        return True
    text = str(raw).strip()
    return not text or bool(NUMBER.match(text))


def _read_sheet(ws, sheet: str, scheme: str, group: str, default_rate,
                out: PlataFile) -> None:
    where = sheet.strip()
    cols = _map_columns(ws, scheme)

    # Ненайденная колонка — не мелочь: часы стали бы нулём, ставка — умолчанием,
    # и расчёт прошёл бы. Поэтому о каждой говорим отдельно и по имени.
    for key in sorted((set(FIELDS) | {"exp_gross"}) - set(cols)):
        out.findings.append(Finding(
            "column", where,
            f"колонка «{_title_of(key, scheme)}» ({key}) не найдена — "
            f"значения этого поля в загрузку не попали",
        ))

    # До самой последней заполненной строки, а не до 19-й: прежний предел
    # молча обрезал листы, где людей больше восемнадцати.
    for r in range(2, (ws.max_row or 1) + 1):
        first, last = ws.cell(r, 2).value, ws.cell(r, 3).value
        named = bool(str(first or "").strip() or str(last or "").strip())

        if not isinstance(ws.cell(r, 1).value, (int, float)):
            if named:
                out.findings.append(Finding(
                    "row", f"{where}, строка {r}",
                    "строка не пронумерована и не загружена: "
                    f"«{str(first or '').strip()} {str(last or '').strip()}»".strip(),
                ))
            continue
        if not first or not last:
            out.findings.append(Finding(
                "row", f"{where}, строка {r}",
                "у строки нет имени или фамилии — она не загружена",
            ))
            continue

        raw = {k: ws.cell(r, c).value for k, c in cols.items()}
        bad = [(key, value) for key, value in raw.items() if not _numeric(value)]
        if bad:
            # Строка целиком не загружается: подставить ноль вместо «восемь»
            # значит посчитать зарплату по числу, которого в таблице нет.
            for key, value in bad:
                out.findings.append(Finding(
                    "value", f"{where}, строка {r}",
                    f"«{value}» в колонке «{_title_of(key, scheme)}» — не число, "
                    f"строка не загружена",
                ))
            continue

        name = f"{str(first).strip()} {str(last).strip()}"
        insured = d(raw.get("insured") or 176)

        out.rows.append(ImportedRow(
            sheet=where,
            scheme=scheme,
            group=group,
            name=name,
            employee=Employee(
                ext_id=name, name=name, group=group, scheme=scheme,
                base_rate=d(raw.get("base_rate") or default_rate),
                coefficient=d(raw.get("coefficient") or 1),
            ),
            timesheet=Timesheet(
                hours={k: d(raw.get(k)) for k in HOUR_FIELDS},
                insured_hours=insured,
                norm_hours=insured,
                deduction=d(raw.get("deduction")),
                cash_payout=d(raw.get("cash")),
                manual_correction=d(raw["correction"]) if raw.get("correction") else None,
            ),
            sheet_meal=d(raw["meal"]) if raw.get("meal") is not None else None,
            expected={
                "net":           d(raw["exp_net"])     if raw.get("exp_net")     else None,
                "gross":         d(raw["exp_gross"])   if raw.get("exp_gross")   else None,
                "contributions": d(raw["exp_contrib"]) if raw.get("exp_contrib") else None,
                "total_cost":    d(raw["exp_total"])   if raw.get("exp_total")   else None,
            },
        ))


def read_plata_file(path: Path | str, default_rate: Decimal | float = 371) -> PlataFile:
    """Разобрать файл: строки и всё, что разобрать не удалось."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = PlataFile()

    # Лист, которого нет в формате, — это либо чужие заметки, либо новая точка,
    # про которую импорт не знает. Второе гораздо хуже, и отличить одно от
    # другого может только человек, поэтому решение отдаём ему.
    for name in wb.sheetnames:
        if name not in SHEET_MAP:
            out.findings.append(Finding(
                "sheet", name,
                "лист не входит в формат PLATA — ни одна его строка не загружена",
            ))

    for sheet, (scheme, group) in SHEET_MAP.items():
        if sheet not in wb.sheetnames:
            out.findings.append(Finding(
                "sheet", sheet.strip(),
                "лист формата не найден в файле — сотрудников с него в загрузке нет",
            ))
            continue
        _read_sheet(wb[sheet], sheet, scheme, group, default_rate, out)

    return out


def read_plata(path: Path | str, default_rate: Decimal | float = 371) -> list[ImportedRow]:
    """Только строки. Оставлено ради сверки движка: ей находки не нужны."""
    return read_plata_file(path, default_rate).rows
