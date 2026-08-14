"""Три выгрузки периода в xlsx (T032): к выплате, для P&L, в виде бухгалтера.

**Все три берут готовый срез** (`reports.sheet.SheetSlice`) — тот самый, что
показан на экране, с тем же значением разреза. Собирать данные заново было бы
проще ровно один раз: дальше две выборки одного и того же расходятся молча, и
человек получает файл, не совпадающий с тем, что он видел. Отсюда же следует
главное свойство безопасности: регистр, которого роль не видит, физически не
доезжает до файла — суммы отобраны политиками базы ещё до среза, а разрез умеет
только сужать.

**Файл опаснее экрана.** Он уходит из продукта и живёт своей жизнью, поэтому
D023 («ни строк, ни следа») здесь читается буквально: в книге не должно быть ни
строк чужого регистра, ни его названия, ни его вклада в итог. Итоги в файле
считаются по его же строкам — не переносятся из полной ведомости.

**Налоги в разрезе регистра не существуют.** Налог и взносы посчитаны по строке
ведомости целиком и живут в `payslip_totals`, закрытой своей политикой (T050).
Поэтому блок налогов есть только в срезе «все видимые регистры»: приписать
общий налог одному регистру значило бы выдумать число.

**И об этом файл говорит вслух** (T141, issue #90). Молчание здесь опаснее
пропуска: файл называется «Строки для P&L», и человек соберёт из него P&L без
налогов на зарплату, ничего не заподозрив. Причины три, и они не смешиваются —
выбран разрез, итоги не отданы роли, налоги не посчитаны вовсе; одна фраза на
все три была бы неправдой в двух случаях из трёх. Про права сказано так, чтобы
не назвать ни сумм, ни людей, ни скрытых регистров: «итоги расчёта вашей роли не
отданы» — факт о правах, верный и тогда, когда в скрытом регистре нет ни одной
строки, поэтому вычесть из него нечего (D023, D014).

Заголовки листов и колонок живут здесь, а не в `web`: выгрузка — это документ,
и его шапка часть формата файла, а не оформления экрана. Единственное, что
приезжает снаружи, — подписи регистров: они уже есть в `web/format.py` и
обязаны совпадать с экраном слово в слово.
"""
from __future__ import annotations

import io
import json
from collections import namedtuple
from datetime import date
from decimal import Decimal
from uuid import UUID

import openpyxl
from django.utils.translation import gettext as _
from django.utils.translation import gettext_noop
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from reports.sheet import ALL, SheetSlice

# Строка блока налогов: статья, точка и два числа. Кортеж, а не словарь: у него
# фиксированный порядок полей, и перепутать местами налог и взносы нельзя.
TaxLine = namedtuple("TaxLine", "article unit tax contributions")

# Строка расхода в файле P&L (T113): статья отчёта, точка, регистр, за что
# именно потрачено и сколько. Тем же кортежем и по тому же доводу, что налоги.
ExpenseLine = namedtuple("ExpenseLine", "article unit ledger title amount")

# Точка у расхода, который ещё не разнесён. Прочерк, а не пустая ячейка: пустая
# читается как «забыли заполнить», прочерк — как «точка ещё не решена». Разница в
# том, будет ли эта сумма кем-то найдена.
NO_UNIT = "—"

# Человек, у которого не заведена статья P&L, обязан быть виден в файле, а не
# пропасть из него: пропавшая строка — это не найденная позже недостача.
#
# Константы с этим текстом больше нет: строка переводится в местах, где
# используется (`gettext("Без статьи")`), потому что перевод возможен только
# когда Django уже настроен, а тело модуля выполняется раньше.

# Колонки таблицы партнёра, под которыми лежит **то же самое**, что у нас.
# Заголовков часов здесь нет намеренно: у партнёра «SATI RADA» — это часы, а у
# нас начисление за них, и деньги под этим заголовком были бы ложью.
PARTNER_HEADERS = {
    "meal_and_vacation_bonus": "TOPLI OBROK I REGRES",
    "minimum_guarantee": "KOREKCIJA DO MINIMALCA",
    "manual_correction": "KOREKCIJA (RUCNO)",
    "deduction": "OBUSTAVA",
}
PARTNER_TOTAL = "UKUPNO ZA ISPLATU"
# Первые две колонки «Вида бухгалтера». Вынесены в константы не ради порядка:
# по ним сверка **узнаёт свою же выгрузку**, когда человек приносит её обратно
# (T119, `reports.own_export`). Заголовки не переводятся намеренно — они и в
# таблице партнёра стоят на сербской латинице, и опознание формата не должно
# зависеть от языка, на котором файл выгрузили.
PARTNER_NUMBER = "R.br."
PARTNER_NAME = "IME I PREZIME"

MONEY = "#,##0.00"


def _stamp(period: date | None) -> str:
    return f"{period:%Y-%m}" if period else "period"


def _file_name(kind: str, period: date | None, cut: str) -> str:
    """Имя файла латиницей: кириллица в заголовке ответа доезжает мусором.

    Разрез входит в имя, когда он выбран: два файла одного месяца в одной папке
    иначе неразличимы, и человек отправит бухгалтеру не тот.
    """
    tail = f"-{cut}" if cut and cut != ALL else ""
    return f"{kind}-{_stamp(period)}{tail}.xlsx"


def _head(ws, text: str, width: int, note: str = "") -> None:
    """Шапка документа: заголовок, при необходимости примечание и пустая строка.

    Примечание стоит **до** строки заголовков колонок, а не среди данных: файл
    собирают в P&L одним разбором, и фраза, положенная строкой таблицы, была бы
    прочитана как позиция с пустой суммой. В преамбуле она читается человеком и
    пропускается загрузчиком — там же, где уже живёт название документа.
    """
    ws.append([text])
    ws["A1"].font = Font(bold=True, size=13)
    if note:
        ws.append([note])
        ws[f"A{ws.max_row}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])
    del width


def _headers(ws, row: list[str]) -> None:
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _fit(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _money_format(ws, first_column: int) -> None:
    """Денежный формат — и одновременно лечение двоичного хвоста float.

    `openpyxl.compat.safe_string` пишет число в XML через `"%.16g" % value` —
    на одну значащую цифру больше, чем нужно для короткого представления
    double, поэтому наружу вылезает вроде `80756.32000000001`. `Decimal` не
    спасает: `%.16g` приводит его к `float` первым делом. Лечится тем, что
    значение кладётся в ячейку строкой, а тип ячейки выставляется вручную —
    тогда `safe_string` отдаёт готовый текст, и в XML попадает ровно то, что
    и должно быть, до копейки.

    Ловушка порядка: обычная запись `cell.value = "80756.32"` через сеттер
    сама переопределит `data_type` на строковый (`_bind_value` смотрит на
    Python-тип), и денежная ячейка стала бы текстом — сумма пропала бы из
    Excel-суммы молча. Поэтому `_value` и `data_type` выставляются в обход
    сеттера, и обе правки (round-trip через строку и сам формат `MONEY`)
    сделаны в одном проходе одной функции: если развести их по разным
    местам, `isinstance` ниже перестанет находить уже подмененные строкой
    ячейки при повторном проходе, и формат проставится не туда или не
    проставится вовсе.

    Округлять значение здесь не нужно и опасно: файл должен показывать ровно
    то, что посчитал движок, а не свою версию с округлением до копейки —
    иначе сумма строк в Excel молча разойдётся с итогом, если где-то в колонке
    окажется величина точнее двух знаков. `str()` у `Decimal` и так отдаёт
    точное представление без округления, поэтому `quantize` не нужен.
    """
    for row in ws.iter_rows(min_col=first_column):
        for cell in row:
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = MONEY
                cell._value = str(Decimal(str(cell.value)))
                cell.data_type = "n"


def _save(book) -> bytes:
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def _named(column, component_title) -> str:
    """Заголовок колонки на языке файла (T092).

    Тот же вопрос, что на экране: в `pay_components.title` лежит подпись,
    замороженная расчётом, а файл выгружает человек, который читает на своём
    языке. Подставляет её вызывающий — здесь, как и с регистрами, только показ.
    """
    return component_title(column.code, column.title) if component_title else column.title


def _titled(title: str, view: SheetSlice, ledger_title) -> str:
    """Шапка документа. Разрез назван, если он выбран, — файл обязан сказать,
    что он не про весь расчёт, иначе его прочитают как полный."""
    if view.cut and view.cut != ALL:
        return f"{title} · {ledger_title(view.cut)}"
    return title


# --- ведомость к выплате ------------------------------------------------------


def payout(view: SheetSlice, *, tenant_id=None, period=None, title="",
           ledger_title=str, component_title=None, item_title=None,
           whole_run=True) -> tuple[bytes, str]:
    """Ведомость к выплате: ровно то, что человек видит на экране, файлом.

    `item_title` и `whole_run` здесь не нужны и не используются: ведомость — про
    зарплату, статьи расхода в ней нет, а налоговой части в ней не бывает ни у
    кого, поэтому и объяснять её отсутствие нечем. Они в подписи потому, что все
    три выгрузки вызываются **одним** набором аргументов (`EXPORTS` в
    `web/reports_views.py` — словарь, а не разбор вида файла): развилка «этому
    виду передаём, тому нет» и есть то место, где однажды передадут не туда.
    """
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "K isplati"

    columns = view.sheet.columns
    _head(ws, _("Ведомость к выплате · %(sub)s") % {"sub": _titled(title, view, ledger_title)}, 4)
    _headers(ws, (
        [_("№"), _("Сотрудник"), _("Точка"), _("Регистр")]
        + [_named(column, component_title) for column in columns]
        + [_("Итого"), _("Примечание")]
    ))

    for number, row in enumerate(view.sheet.rows, start=1):
        notes = []
        if row.is_retro:
            # Разница за закрытый месяц обязана объяснить себя: без источника
            # это непонятная сумма в чужом месяце (T026).
            notes.append(_("разница за %(date)s") % {"date": f"{row.retro_source:%m.%Y}"})
        if row.frozen:
            notes.append(
                _("строка заморожена: %(reason)s") % {"reason": row.freeze_reason}
                if row.freeze_reason else _("строка заморожена")
            )
        ws.append(
            [number, row.employee, row.unit, ledger_title(row.ledger)]
            + [row.amounts.get(column.code) for column in columns]
            + [row.total, "; ".join(notes)]
        )

    # Подвал считается по строкам этого файла, а не переносится из полной
    # ведомости: итог, больший суммы показанного, выдаёт скрытое вычитанием.
    _headers(ws, (
        [_("Итого"), "", "", ""]
        + [view.sheet.column_totals.get(column.code) for column in columns]
        + [view.sheet.total, _("человек: %(count)s") % {"count": view.sheet.employees}]
    ))

    _fit(ws, [5, 28, 8, 15] + [14] * len(columns) + [14, 30])
    _money_format(ws, 5)
    return _save(book), _file_name("payout", period, view.cut)


# --- строки для P&L -----------------------------------------------------------

# Почему в файле нет налоговой части. Кодом, а не готовой фразой: код называет
# положение дел, а слова к нему подбираются там, где известен язык читателя, —
# и один и тот же код объясняет файл и экран одинаково (T141).
TAXES_BY_CUT = "cut"            # выбран разрез: налог не принадлежит регистру
TAXES_WITHHELD = "withheld"     # итоги расчёта роли не отданы
TAXES_NOT_COUNTED = "absent"    # налогов в расчёте этого месяца нет вовсе

# Слова к причинам. Здесь, а не в `web`: это часть документа, как и заголовки
# колонок, — файл уходит из продукта и обязан объяснять себя сам. Экран берёт
# те же строки отсюда: две формулировки одного и того же разъехались бы молча.
TAXES_MISSING_TEXTS = {
    TAXES_BY_CUT: gettext_noop(
        "Налоговая часть в этот файл не вошла: налог и взносы считаются по "
        "строке ведомости целиком, и разрезу по регистру они не принадлежат. "
        "Для P&L с налогами возьмите файл без разреза."
    ),
    # Ни сумм, ни имён, ни названия регистра: это факт о правах, а не о данных
    # (D023). Он верен и тогда, когда скрывать нечего, — вычесть из него нельзя.
    TAXES_WITHHELD: gettext_noop(
        "Налоговая часть в этот файл не вошла: налог и взносы считаются по "
        "строке ведомости целиком, а итоги расчёта вашей роли не отданы. "
        "Собранный по этому файлу P&L окажется без налогов на зарплату — "
        "возьмите файл у того, кому отдан весь расчёт."
    ),
    TAXES_NOT_COUNTED: gettext_noop(
        "Налоговая часть в этот файл не вошла: налог и взносы в расчёте этого "
        "месяца не посчитаны."
    ),
}


def taxes_missing(*, has_taxes: bool, has_rows: bool, cut: str, whole_run: bool) -> str:
    """Код причины, по которой в файле нет налогов, или пусто, если всё на месте.

    Порядок проверок не произвольный.

    Сначала «есть ли о чём говорить»: у месяца без единой зарплатной строки
    налоговой части не ожидают, и надпись про неё была бы шумом.

    **Дальше разрез — раньше вопроса о том, есть ли налоги вообще.** В разрез
    налог не попадает никогда, сколько бы его ни было посчитано (`pnl` обнуляет
    блок налогов на любом разрезе), поэтому «налоги у периода есть» тут ничего
    не решает. Проверено это не рассуждением: пока вопрос стоял первым, экран
    директора с выбранным разрезом молчал, а файл того же разреза приезжал без
    налогов и объяснял это надписью — то самое расхождение экрана и файла,
    против которого написана вся эта задача (найдено смоуком на стенде).

    И только потом права: у роли, которой отдан весь расчёт, остаётся
    единственная правда — налогов не посчитали.
    """
    if not has_rows:
        return ""
    if cut != ALL:
        return TAXES_BY_CUT
    if has_taxes:
        return ""
    return TAXES_NOT_COUNTED if whole_run else TAXES_WITHHELD


def taxes_note(reason: str) -> str:
    """Причина словами читателя. Пустая причина — пустая надпись."""
    known = TAXES_MISSING_TEXTS.get(reason)
    return _(known) if known else ""


def collect_articles(tenant_id: UUID) -> dict[str, str]:
    """Ключ сотрудника → статья P&L его группы.

    Через условия найма, а не через имя листа: соответствие «лист → точка»
    приблизительное, а группа у человека одна и версионируется вместе с наймом.
    """
    from core.models import EmploymentTerm

    out: dict[str, str] = {}
    for term in EmploymentTerm.objects.filter(
        tenant_id=tenant_id
    ).select_related("employee", "group__pnl_item").order_by("valid_from"):
        if term.group.pnl_item_id:
            out[term.employee.external_id] = term.group.pnl_item.title
    return out


def collect_taxes(tenant_id: UUID, period: date, articles: dict[str, str]) -> list[TaxLine]:
    """Налог и взносы по статье и точке — из итогов строк ведомости.

    Выборка идёт от `payslip_totals`, и это и есть граница видимости: итоги
    видны роли, только если ей видна вся строка (T050). Приложение здесь ничего
    не маскирует и не досчитывает — что база отдала, то и сложено.
    """
    from core.models import PayslipTotals

    grouped: dict[tuple[str, str], list[Decimal]] = {}
    for row in PayslipTotals.objects.filter(
        tenant_id=tenant_id, payslip__payrun__period=period
    ).select_related("payslip__employee", "payslip__unit"):
        key = (
            articles.get(row.payslip.employee.external_id, _("Без статьи")),
            row.payslip.unit.code if row.payslip.unit_id else "",
        )
        bucket = grouped.setdefault(key, [Decimal(0), Decimal(0)])
        bucket[0] += row.tax
        bucket[1] += row.contributions

    return [
        TaxLine(article, unit, tax, contributions)
        for (article, unit), (tax, contributions) in sorted(grouped.items())
        if tax or contributions
    ]


# Что считается строкой расхода периода. Одним куском, потому что спрашивают его
# дважды: сами строки (`collect_expenses`) и набор их регистров (`fact_ledgers`,
# из него собирается переключатель разрезов). Разъехавшись, эти два условия дали
# бы разрез, которому нечего сужать, — или наоборот, спрятанный разрез с
# деньгами внутри (T137, issue #108). Пользовательских значений здесь нет:
# кусок подставляется в запрос как есть, а данные по-прежнему идут параметрами.
EXPENSE_LINES = "l.kind <> 'transfer' and l.source <> 'payroll'"


def fact_ledgers(tenant_id: UUID, period: date) -> list[str]:
    """Регистры, в которых у периода есть расходы (T137).

    Нужны переключателю разрезов: до этого он знал только регистры ведомости, и
    трата в регистре без зарплаты разрезом не отделялась (issue #108).

    Срез делает база: `pnl_lines` объявлено `security_invoker`, поэтому регистр,
    которого роль не видит, отсюда не приезжает — а значит и кнопки с его
    названием на экране не появляется (D023).
    """
    from django.db import connection

    with connection.cursor() as cursor:
        cursor.execute(
            f"""select distinct l.ledger::text
                  from pnl_lines l
                 where l.tenant_id = %s and l.period = %s and {EXPENSE_LINES}""",
            [str(tenant_id), period],
        )
        return [row[0] for row in cursor.fetchall()]


def taxes_exist(tenant_id: UUID, period: date) -> bool:
    """Есть ли у периода налоговая часть, видимая смотрящему (T141).

    Спрашивается **тем же сбором**, который кладёт налоги в файл, а не отдельной
    выборкой рядом: экран говорит человеку, чего в файле не будет, и ответить на
    этот вопрос по-разному они не имеют права. Две выборки одного и того же
    расходятся молча — и тогда экран обещает налоги, а файл приезжает без них.
    """
    return bool(collect_taxes(tenant_id, period, {}))


def collect_expenses(tenant_id: UUID, period: date, cut: str,
                     item_title=None) -> list[ExpenseLine]:
    """Расходы периода строками для P&L — из представления `pnl_lines` (T113).

    **Читается представление, а не таблица.** В нём уже записано, что такое
    строка P&L: действующая версия (не заменённая) и не родитель разнесения
    (`split` исключён, вместо него в отчёт идут дети по точкам). Повторить эти
    два условия выборкой здесь было бы третьей копией одного правила в этом
    блоке — а две предыдущие разъехались молча (`allocate_fact` терял статью,
    `allocation_plan` строил план по половине точек).

    **Срез делает база.** Ни одного условия про права здесь нет: `pnl_lines`
    объявлено `security_invoker`, то есть политики `facts` действуют внутри него
    (D014). Разрез по регистру только **сужает** видимое.

    **Зарплата отсюда не берётся.** Зарплатная половина файла приезжает из
    ведомости — того самого среза, который человек видел на экране. Факты с
    источником `payroll` пропускаются, и это не осторожность на всякий случай:
    проводка зарплаты в факты отложена (журнал блока `facts`, T107 и T113), и в
    день, когда её сделают, файл не должен начать считать одни и те же деньги
    дважды.

    Переводы (`kind = 'transfer'`) пропускаются по той же причине, по какой их
    не считают `pnl_by_unit` и `pnl_by_network`: пополнение кассы — не расход.
    """
    from django.db import connection

    with connection.cursor() as cursor:
        cursor.execute(
            f"""select l.pnl_title,
                      coalesce(l.unit_code, ''),
                      l.ledger::text,
                      l.title,
                      -- `::text` не для красоты: Django ставит psycopg свой
                      -- загрузчик jsonb, который отдаёт сырую строку (разбирать
                      -- её — дело `JSONField`, а его тут нет). Сырым курсором
                      -- значение приходит то строкой, то словарём в зависимости
                      -- от того, кто настроил соединение; явное приведение
                      -- делает ответ одним всегда.
                      e.titles::text,
                      sum(l.amount)
                 from pnl_lines l
                 left join expense_items e on e.id = l.expense_item_id
                where l.tenant_id = %s
                  and l.period = %s
                  and {EXPENSE_LINES}
                  and (%s = '' or l.ledger::text = %s)
                group by 1, 2, 3, 4, 5
                order by 1, 2, 4""",
            [str(tenant_id), period, cut, cut],
        )
        found = cursor.fetchall()

    return [
        ExpenseLine(
            article, unit, ledger,
            # Название на языке файла — из статьи; её нет (расход не из кассы,
            # а, скажем, фактура) — остаётся снимок, замороженный записью.
            (item_title(json.loads(titles)) if item_title and titles else "") or snapshot,
            amount,
        )
        for article, unit, ledger, snapshot, titles, amount in found
    ]


def pnl(view: SheetSlice, *, tenant_id=None, period=None, title="",
        ledger_title=str, articles=None, taxes=None, expenses=None,
        component_title=None, item_title=None, whole_run=True) -> tuple[bytes, str]:
    """Строки для P&L: начисления и налоги раздельно, по статье и точке.

    Итоговой строки в файле нет намеренно: это заготовка строк для сборки P&L,
    и «Итого» в ней загрузчик посчитал бы такой же строкой данных.

    `whole_run` — отдан ли смотрящему весь расчёт партнёра. Приезжает снаружи, а
    не спрашивается здесь: вопрос о правах задают функции, на которых стоят
    политики базы (`web/runslice.py`), и второй правды о видимости рядом с
    первой быть не должно (D014). Умолчание `True` — «прав достаточно»: тогда
    отсутствие налогов не объясняется правами наугад.
    """
    if articles is None:
        articles = collect_articles(tenant_id)
    if expenses is None:
        # Разрез сужает и расходы тоже: файл одного регистра, где зарплата
        # своего регистра, а траты всех, не сходится ни с чем.
        expenses = collect_expenses(tenant_id, period, view.cut, item_title)
    # Разрез — это один регистр, а налог посчитан по строке ведомости целиком.
    # Приписать его регистру нельзя, поэтому в разрезе налогов нет вовсе — и это
    # решается здесь, а не тем, что вызывающий их не передал.
    if view.cut != ALL:
        taxes = []
    elif taxes is None:
        taxes = collect_taxes(tenant_id, period, articles)

    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "PnL"

    # Чего в файле нет — сказано в шапке, а не оставлено человеку на догадку
    # (T141): файл называется «Строки для P&L», и P&L по нему соберут как есть.
    missing = taxes_missing(
        has_taxes=bool(taxes), has_rows=bool(view.sheet.rows),
        cut=view.cut, whole_run=whole_run,
    )
    _head(
        ws, _("Строки для P&L · %(sub)s") % {"sub": _titled(title, view, ledger_title)},
        4, taxes_note(missing),
    )
    _headers(ws, [
        _("Статья P&L"), _("Точка"), _("Регистр"), _("Тип строки"), _("Компонент"), _("Сумма"),
    ])

    accruals: dict[tuple[str, str, str, str, str], Decimal] = {}
    for row in view.sheet.rows:
        # Ключом, а не отображаемым именем: справочник собран по
        # `employees.external_id`, и спрошенный именем он не отвечал никогда —
        # у всех начислений стояло «Без статьи» (issue #95). Налоговые строки
        # спрашивали правильно, поэтому дефект и выглядел как «статья есть,
        # но не у всех».
        article = articles.get(row.employee_key or row.employee, _("Без статьи"))
        for column in view.sheet.columns:
            amount = row.amounts.get(column.code)
            if not amount:
                continue
            key = (article, row.unit, row.ledger, column.code,
                   _named(column, component_title))
            accruals[key] = accruals.get(key, Decimal(0)) + amount

    for (article, unit, ledger, _code, column_title), amount in sorted(accruals.items()):
        ws.append([article, unit, ledger_title(ledger), _("Начисление"), column_title, amount])

    # Расходы — рядом с начислениями и в тех же статьях, а не отдельным листом:
    # файл собирают в P&L одним разбором, и вторая таблица со своими колонками
    # для этого бесполезна ровно так же, как её отсутствие.
    spent: dict[tuple[str, str, str, str], Decimal] = {}
    for line in expenses:
        key = (line.article, line.unit or NO_UNIT, line.ledger, line.title)
        spent[key] = spent.get(key, Decimal(0)) + line.amount
    for (article, unit, ledger, spent_on), amount in sorted(spent.items()):
        ws.append([article, unit, ledger_title(ledger), _("Расход"), spent_on, amount])

    for line in taxes:
        # Регистра у налога нет, и прочерк здесь честнее пустой ячейки: пустая
        # читается как «забыли заполнить».
        ws.append([line.article, line.unit, "—", _("Налог"), _("Налог на доход"), line.tax])
        ws.append([line.article, line.unit, "—", _("Взносы"), _("Взносы"), line.contributions])

    _fit(ws, [26, 8, 15, 14, 26, 14])
    _money_format(ws, 6)
    return _save(book), _file_name("pnl", period, view.cut)


# --- вид, привычный бухгалтеру ------------------------------------------------


def partner(view: SheetSlice, *, tenant_id=None, period=None, title="",
            ledger_title=str, component_title=None,
            item_title=None, whole_run=True) -> tuple[bytes, str]:
    """Тот же расчёт, разложенный так, как привык читать бухгалтер партнёра.

    Привычное здесь — две вещи: **лист на точку** (в его таблице лист на точку
    и схему) и **его собственные заголовки** там, где под ними лежит то же
    самое. Где не то же самое, остаётся наше название: заголовок «SATI RADA»
    над начислением за часы был бы не привычным видом, а неправдой.
    """
    book = openpyxl.Workbook()
    book.remove(book.active)

    columns = view.sheet.columns
    headers = [
        PARTNER_HEADERS.get(column.code) or _named(column, component_title)
        for column in columns
    ]

    units = sorted({row.unit for row in view.sheet.rows})
    for unit in units or [""]:
        ws = book.create_sheet(unit or "Bez objekta")
        _head(ws, f"{_titled(title, view, ledger_title)} · {unit}", 4)
        _headers(
            ws,
            [PARTNER_NUMBER, PARTNER_NAME, _("Регистр")] + headers + [PARTNER_TOTAL],
        )

        rows = [row for row in view.sheet.rows if row.unit == unit]
        for number, row in enumerate(rows, start=1):
            ws.append(
                [number, row.employee, ledger_title(row.ledger)]
                + [row.amounts.get(column.code) for column in columns]
                + [row.total]
            )

        # Подвал листа — по строкам этого листа. Складывать сюда весь период
        # значило бы показать на листе точки чужие деньги.
        _headers(ws, (
            ["UKUPNO", "", ""]
            + [
                sum((row.amounts.get(column.code) or Decimal(0) for row in rows), Decimal(0))
                or None
                for column in columns
            ]
            + [sum((row.total for row in rows), Decimal(0))]
        ))
        _fit(ws, [6, 28, 15] + [16] * len(columns) + [16])
        _money_format(ws, 4)

    return _save(book), _file_name("partner", period, view.cut)
