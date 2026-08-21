"""
Что показывает наполненная демо-база: продукт целиком, а не пустые экраны.

Definition of Done блока говорит прямо: демо готово, когда **в нём работают и
отчёт расхождений, и сверка**. Это не про то, что код отчёта не падает, — это
про то, что на демо-данных ему есть что показать. Отчёт расхождений без второго
месяца и сверка без файла честно пусты, и такое демо показывать нельзя.

Проверки идут на настоящей базе, наполненной той же командой, которую запускает
человек. Отчёты считаются в подпроцессе через `manage.py shell`: настройки
Django читаются один раз на процесс, а в прогоне тестов уже поднята другая база.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from decimal import Decimal

import pytest

from conftest import MANAGE_PY, temp_database
from demo.dataset import MONTHS

CYRILLIC = re.compile(r"[а-яА-ЯёЁ]")


def manage(dsn: str, *args: str) -> subprocess.CompletedProcess:
    env = {
        **os.environ,
        "DATABASE_URL": dsn,
        "DEMO_DATABASE_URL": dsn,
        "SECRET_KEY": "test-only-not-a-secret",
        "DJANGO_SETTINGS_MODULE": "config.settings",
    }
    return subprocess.run(
        [sys.executable, str(MANAGE_PY), *args],
        env=env, capture_output=True, text=True, check=True,
    )


@pytest.fixture(scope="module")
def demo_db():
    with temp_database("content") as dsn:
        manage(dsn, "seed_demo")
        yield dsn


@pytest.fixture
def conn(demo_db):
    psycopg = pytest.importorskip("psycopg")

    from core.db_types import register_enum_types

    with psycopg.connect(demo_db) as c:
        register_enum_types(c)
        yield c


def report(dsn: str, script: str) -> dict:
    """Посчитать что-нибудь внутри Django и вернуть результат словарём.

    Через `manage.py shell`, потому что отчёты живут в ORM, а второй базы в одном
    процессе Django не бывает. Печатается одна строка JSON — её и читаем.
    """
    out = manage(dsn, "shell", "-c", script).stdout
    return json.loads(out.strip().splitlines()[-1])


# --- организация ---------------------------------------------------------------


def test_two_legal_entities_three_units_thirty_people(conn):
    tenant = conn.execute("select id, title from tenants").fetchall()
    assert len(tenant) == 1, "в демо-базе должен быть ровно один партнёр"

    assert conn.execute("select count(*) from legal_entities").fetchone()[0] == 2
    units = [
        row[0]
        for row in conn.execute("select code from units order by code").fetchall()
    ]
    assert units == ["BG1", "NS1", "NS2"]
    # Обе точки Нови-Сада — на одном юрлице, Белград — на другом: два юрлица без
    # разделения точек не показывают ничего.
    per_entity = conn.execute(
        "select legal_entity_id, count(*) from units group by 1 order by 2"
    ).fetchall()
    assert [row[1] for row in per_entity] == [1, 2]

    assert conn.execute("select count(*) from employees").fetchone()[0] == 30


def test_three_months_two_of_them_closed(conn):
    periods = conn.execute(
        "select period, status from periods order by period"
    ).fetchall()
    assert len(periods) == 3
    assert [str(row[1]) for row in periods] == ["closed", "closed", "open"]

    payruns = conn.execute(
        "select period, status from payruns order by period"
    ).fetchall()
    assert [str(row[1]) for row in payruns] == ["approved", "approved", "calculated"]


def test_every_ledger_and_every_scheme_is_represented(conn):
    ledgers = {
        row[0]
        for row in conn.execute("select distinct ledger from pay_components").fetchall()
    }
    assert ledgers == {"official", "supplementary", "internal"}

    schemes = {
        row[0]
        for row in conn.execute(
            """select distinct coalesce(t.scheme, g.scheme)
                 from employment_terms t join employee_groups g on g.id = t.group_id"""
        ).fetchall()
    }
    assert {"standard", "half_time", "half_time_min_base", "temporary"} <= schemes


def test_employment_terms_are_versioned(conn):
    """У кого-то две версии условий найма: версионирование должно быть видно."""
    versions = conn.execute(
        """select employee_id, count(*) from employment_terms
            group by 1 having count(*) > 1"""
    ).fetchall()
    assert versions, "ни у кого нет второй версии условий найма"


def test_manual_correction_has_an_author_and_a_reason(conn):
    corrections = conn.execute(
        """select manual_correction, correction_reason, corrected_by
             from timesheets where manual_correction is not null"""
    ).fetchall()
    assert corrections
    for amount, reason, author in corrections:
        assert amount is not None and reason and author, "правка без следа (D025)"


# --- английский ----------------------------------------------------------------


def test_nothing_a_visitor_sees_is_written_in_russian(conn):
    """Демо всегда англоязычное — во всём, что лежит в базе строкой.

    Подписи компонентов выплаты сюда **не входят намеренно** (T092). В
    `pay_components.title` лежит подпись, замороженная расчётом на языке правил;
    посетителю она не показывается — колонка называется по правилам на языке
    страницы. Проверять здесь русское слово значило бы требовать от записи в
    базе того, чем она не является, и это требование однажды починили бы, сломав
    объяснение закрытого месяца. Что видит посетитель, проверяет тест ниже.
    """
    seen: list[str] = []
    for sql in (
        "select title from tenants",
        "select title from legal_entities",
        "select title from units",
        "select title from employee_groups",
        "select title from roles",
        "select first_name from employees",
        "select last_name from employees",
        "select correction_reason from timesheets where correction_reason is not null",
        # Расходы (T113): и справочник статей, и сами траты. Названия статей —
        # это то, что посетитель видит в форме внесения и в выгрузке P&L, а
        # `facts.title` уезжает в файл как «за что деньги».
        "select jsonb_each_text.value from expense_items, jsonb_each_text(titles)",
        "select title from facts",
        "select note from facts where note is not null",
    ):
        seen += [row[0] for row in conn.execute(sql).fetchall() if row[0]]

    russian = sorted({text for text in seen if CYRILLIC.search(text)})
    assert not russian, f"по-русски в демо: {russian}"


def test_the_columns_a_visitor_sees_are_english(demo_db):
    """Колонки ведомости демо названы по-английски — тем же путём, что у партнёра.

    Спрашивается ровно то, что подставит страница: подписи компонентов на языке
    демо (`UI_LANGUAGE=en`). До T092 демо держало свой список английских подписей
    и выглядело англоязычным даже тогда, когда продукт показывал колонки
    по-русски. Список убран, и эта проверка — то, что стоит на его месте.
    """
    result = report(demo_db, """
import json
from datetime import date
from django.utils import translation
from core.models import Tenant
from web.labels import component_titles
t = Tenant.objects.get(code="demo")
with translation.override("en"):
    titles = component_titles(t.id, t.country_code, date(2026, 6, 1))
print(json.dumps(titles, ensure_ascii=False))
""")
    assert result, "подписей компонентов нет вовсе — колонки останутся без названий"
    russian = sorted(v for v in result.values() if CYRILLIC.search(v))
    assert not russian, f"колонки демо по-русски: {russian}"
    assert result.get("hours.regular") == "Worked hours", result.get("hours.regular")


# --- отчёты, ради которых стенд и наполняли ------------------------------------


def test_variance_report_has_something_to_show(demo_db):
    """Отчёт расхождений содержателен в обоих переходах, а не «сравнивать нечего».

    Ровно поэтому в демо три месяца, а не два: один переход показал бы одну
    пару, и пустой отчёт на второй паре списали бы на «так и должно быть».
    """
    result = report(demo_db, """
import json
from datetime import date
from core.models import Tenant
from reports.variance import build_variance
t = Tenant.objects.get(code="demo")
out = {}
for label, p in (("july", date(2026, 7, 1)), ("august", date(2026, 8, 1))):
    r = build_variance(t.id, p)
    out[label] = {
        "lines": len(r.lines), "compared": r.compared,
        "employees": r.employees, "nothing": r.nothing_to_compare,
    }
print(json.dumps(out))
""")
    for label, body in result.items():
        assert not body["nothing"], f"{label}: сравнивать не с чем"
        assert body["compared"] > 0
        assert body["lines"] >= 3, f"{label}: в отчёте почти пусто ({body['lines']})"
        assert body["employees"] >= 2, f"{label}: отклонения у одного человека"


def test_reconciliation_shows_all_of_its_states(demo_db):
    """Сверка на демо-файле показывает и совпадение, и все три вида расхождения.

    Файл собирается из самих демо-данных и несёт нарочные расхождения (см.
    `demo.table`). Проверяется не «сверка не падает», а что посетитель видит на
    экране все её состояния: сошлось, разошлось с объяснением, копейки, и люди,
    которых нет с одной из сторон.
    """
    result = report(demo_db, """
import io, json
from datetime import date
from core.models import Tenant
from demo.accountant_table import build_accountant_table
from demo.seed import det_id
from demo.table import accountant_rows
from reports.reconcile import reconcile
from web.dbcontext import db_context
period = date(2026, 6, 1)
book = build_accountant_table(accountant_rows(period))
t = Tenant.objects.get(code="demo")
# Сверка спрашивает базу, отдан ли роли расчёт целиком (T100), поэтому
# зовётся с выставленным контекстом — тем же `db_context`, каким ходит
# фоновая задача. Без контекста ответ базы «расчёта вам не отдали», и
# сверка молчит о нём: это правильное поведение (D014), но проверять на
# нём состояния экрана нельзя — посетитель демо всегда кем-то вошёл.
with db_context(det_id("user", "director")):
    r = reconcile(io.BytesIO(book), tenant_id=t.id, period=period)
print(json.dumps({
    "matched": sum(1 for line in r.lines if line.matched),
    "off": sum(1 for line in r.lines
               if line.compared and not line.matched and not line.rounding_only),
    "rounding": sum(1 for line in r.lines if line.rounding_only),
    "only_in_file": len(r.only_in_file),
    "only_in_run": len(r.only_in_run),
    "findings": len(r.findings),
    "causes": sum(len(line.causes) for line in r.lines),
}))
""")
    assert result["matched"] >= 20, "почти ничего не сошлось — файл не из этих данных"
    assert result["off"] >= 1, "нет ни одного существенного расхождения"
    assert result["rounding"] >= 1, "нет расхождения-округления"
    assert result["only_in_file"] >= 1, "нет человека, оставшегося только в файле"
    assert result["only_in_run"] >= 1, "нет человека, которого нет в таблице"
    assert result["causes"] >= 1, "расхождение не объясняет себя входами"
    assert result["findings"] == 0, "демо-файл разобрался с находками — формат поехал"


def test_payroll_sheet_is_not_empty_for_a_closed_month(demo_db):
    """Ведомость закрытого месяца показывает строки и колонки, а не пустоту."""
    result = report(demo_db, """
import json
from datetime import date
from core.models import Tenant
from reports.sheet import build_slice
t = Tenant.objects.get(code="demo")
s = build_slice(t.id, date(2026, 6, 1))
print(json.dumps({
    "rows": len(s.sheet.rows),
    "columns": len(s.sheet.columns),
    "ledgers": sorted(cut.code for cut in s.cuts),
}))
""")
    assert result["rows"] >= 29
    assert result["columns"] >= 3
    assert len(result["ledgers"]) >= 3, "разрезов по регистрам меньше трёх"


# --- расходы из кассы (T113) ---------------------------------------------------
#
# Демо без расходов показывает половину продукта и молчит об этом: P&L
# складывается из зарплаты и трат, и вторая половина у партнёра как раз та,
# которую сегодня никто не собирает.


def test_every_month_has_cash_expenses(conn):
    """Траты есть в каждом месяце, включая закрытые.

    Закрытые важнее открытого: именно по ним посетитель открывает выгрузку и
    видит, что в ней обе части. Положить их туда можно только пока месяц
    открыт — поэтому наполнение пишет траты до утверждения расчёта, а не после.
    """
    by_month = dict(conn.execute(
        """select to_char(period, 'YYYY-MM'), count(*)
             from facts where source = 'manual' and superseded_at is null
            group by 1 order by 1"""
    ).fetchall())
    assert set(by_month) == {"2026-06", "2026-07", "2026-08"}, by_month
    assert all(count > 0 for count in by_month.values()), by_month


def test_an_expense_of_the_whole_network_is_spread_to_the_kopeck(conn):
    """Аренда офиса разошлась по трём точкам, и сумма детей равна родителю."""
    parent, children = conn.execute(
        """select f.amount,
                  (select coalesce(sum(c.amount), 0) from facts c
                    where c.parent_fact_id = f.id and c.superseded_at is null)
             from facts f
             join expense_items e on e.id = f.expense_item_id
            where e.code = 'office_rent' and f.parent_fact_id is null
              and f.period = '2026-06-01' and f.superseded_at is null"""
    ).fetchone()
    assert parent == children, f"разнесение потеряло деньги: {parent} против {children}"

    units = conn.execute(
        """select count(distinct c.unit_id) from facts c
             join facts f on f.id = c.parent_fact_id
             join expense_items e on e.id = f.expense_item_id
            where e.code = 'office_rent' and c.period = '2026-06-01'"""
    ).fetchone()[0]
    assert units == 3, f"аренда легла не на все точки: {units}"


def test_something_is_waiting_to_be_allocated(conn):
    """Нераспределённая сумма в демо есть — и она видна там, где её ищут.

    Состояние «сумма есть, точка ещё не решена» — не поломка, а ответ продукта,
    и показать его надо. Демо, где всё разнесено, обещает, что так бывает
    всегда.
    """
    waiting = conn.execute(
        "select title, amount from facts_unallocated order by amount desc"
    ).fetchall()
    assert waiting, "в демо нечему быть нераспределённым — экран будет пустым"
    assert any(amount > 0 for _title, amount in waiting), waiting


def test_the_pnl_export_of_a_closed_month_has_both_halves(demo_db):
    """Выгрузка «Строки для P&L» содержит и зарплату, и траты.

    Это и есть Definition of Done третьей очереди: «расходы попадают в выгрузку
    рядом с зарплатными строками, в тех же статьях». Проверяется на **закрытом**
    месяце — том, который бухгалтер отдаёт дальше.
    """
    result = report(demo_db, """
import base64, io, json
from datetime import date
from django.utils import translation
from core.models import Tenant
from reports.export import pnl
from reports.sheet import build_slice
from web.cash import item_title
t = Tenant.objects.get(code="demo")
with translation.override("en"):
    body, name = pnl(
        build_slice(t.id, date(2026, 6, 1)), tenant_id=t.id, period=date(2026, 6, 1),
        title="June 2026", item_title=item_title,
    )
import openpyxl
sheet = openpyxl.load_workbook(io.BytesIO(body)).active
rows = [r for r in sheet.iter_rows(values_only=True) if r and r[3]]
print(json.dumps({"kinds": sorted({r[3] for r in rows}),
                  "articles": sorted({r[0] for r in rows if r[3] == "Expense"})},
                 ensure_ascii=False))
""")
    assert "Accrual" in result["kinds"], result
    assert "Expense" in result["kinds"], f"в выгрузке демо нет расходов: {result}"
    assert result["articles"], result
    russian = sorted(a for a in result["articles"] if CYRILLIC.search(a))
    assert not russian, f"статьи расходов в выгрузке демо по-русски: {russian}"


# --- касса и НДС (T145, T146) ---------------------------------------------------


def test_the_demo_shows_tills_and_the_ledger_that_follows_them(conn):
    """Демо показывает кассы и то, ради чего они заведены (D039).

    Не «поле есть», а правило видно: у одной точки две кассы разных регистров, и
    расход из внутренней кассы лежит во внутреннем регистре. С одной кассой на
    точку это выглядело бы лишним полем в форме.
    """
    tills = conn.execute(
        "select t.code, u.code, t.ledger::text from tills t join units u on u.id = t.unit_id"
    ).fetchall()
    assert len(tills) >= 4, tills

    by_unit = {}
    for _code, unit, ledger in tills:
        by_unit.setdefault(unit, []).append(ledger)
    assert any(len(ledgers) > 1 for ledgers in by_unit.values()), (
        "у всех точек по одной кассе — правило «регистр следует из кассы» не показано"
    )

    paid = conn.execute(
        """select f.ledger::text, t.ledger::text
             from facts f join tills t on t.id = f.till_id
            where f.superseded_at is null"""
    ).fetchall()
    assert paid, "ни один расход демо не оплачен из кассы"
    assert all(fact == till for fact, till in paid), paid


def test_the_demo_has_expenses_with_vat_and_without(conn):
    """В демо есть и траты с выделенным налогом, и без него (D042).

    Оба состояния законные, и оба должны быть видны: с налогом — чтобы в P&L
    было видно разницу между суммой документа и суммой без НДС, без него — чтобы
    было видно, что пустая ставка не превращает сумму в ноль.
    """
    rows = conn.execute(
        """select count(*) filter (where vat_rate is not null),
                  count(*) filter (where vat_rate is null),
                  coalesce(sum(vat_amount), 0)
             from facts where superseded_at is null and source = 'manual'"""
    ).fetchone()
    with_vat, without_vat, total_vat = rows
    assert with_vat > 0, "в демо нет ни одной траты с НДС"
    assert without_vat > 0, "в демо все траты с НДС — состояние «налог не выделен» не показано"
    assert total_vat > 0, "ставка есть, а сумма налога не посчиталась"


def test_the_net_amount_is_smaller_exactly_by_the_vat(conn):
    """Сумма без НДС отличается от суммы документа ровно на налог, и не иначе."""
    off = conn.execute(
        """select count(*) from pnl_lines
            where amount_net <> amount - coalesce(vat_amount, 0)"""
    ).fetchone()[0]
    assert off == 0, f"{off} строк, где сумма без НДС не сходится с суммой документа"


def test_the_month_a_visitor_lands_on_has_till_and_vat_facts(conn):
    """Экран, открывающийся посетителю сразу, показывает кассу и НДС (Н7, сверка 8).

    Демо приводит посетителя на **открытый** месяц — последний в
    `demo.dataset.MONTHS`, а не на приколоченную строку даты: сид пересобирается,
    и открытый месяц у него всегда последний. Если в этом месяце ни у одного
    факта нет `till_id` или `vat_rate`, посетитель, ничего не меняя в фильтре
    дат, видит две пустые колонки — то самое, что нашла сверка.
    """
    open_period = MONTHS[-1].period
    row = conn.execute(
        """select count(*) filter (where till_id is not null),
                  count(*) filter (where vat_rate is not null)
             from facts
            where period = %s and superseded_at is null""",
        [open_period],
    ).fetchone()
    with_till, with_vat = row
    assert with_till > 0, f"в {open_period:%Y-%m} нет ни одного расхода из кассы"
    assert with_vat > 0, f"в {open_period:%Y-%m} нет ни одного расхода с НДС"


# --- поставщики: контрагенты, счета, платежи (T151-T153) -----------------------
#
# Definition of Done блока suppliers прямо про демо: оно обязано показывать
# контрагентов, неоплаченный счёт и непустой инбокс. Без этого раздела демо не
# знает о четвёртой очереди вовсе.


def test_the_demo_has_counterparties_with_an_empty_dodo_is_key(conn):
    """Контрагенты в демо есть, и ключ Dodo IS у них пуст (T150).

    Пусто — не недосмотр: сведение со справочником поставщиков Dodo IS
    случится только в шестой очереди. Заполненное поле сегодня соврало бы про
    то, что оно уже случилось.
    """
    rows = conn.execute("select title, external_id from counterparties").fetchall()
    assert len(rows) >= 4, "в демо меньше контрагентов, чем описано в задаче"
    assert all(external_id is None for _title, external_id in rows), (
        f"у контрагента демо уже стоит ключ Dodo IS: {rows}"
    )


def _invoice_balances(conn) -> list[tuple]:
    """Сумма документа и сумма оплат по нему — по каждому счёту поставщика.

    Остаток не хранится колонкой: колонки «оплачено» у счёта нет намеренно
    (`web/suppliers.py`), он считается сложением платежей, и запрос повторяет
    именно эту логику, а не читает выдуманный флаг.
    """
    return conn.execute(
        """select d.total_amount, coalesce(p.paid, 0)
             from source_documents d
             left join lateral (
                 select sum(f.amount) as paid
                   from facts f
                  where f.document_id = d.id
                    and f.dedup_key like 'manual:payment:%'
                    and f.superseded_at is null
             ) p on true
            where d.kind = 'invoice' and d.source = 'manual'"""
    ).fetchall()


def test_there_is_an_invoice_that_was_never_paid(conn):
    """Есть счёт без единой оплаты: остаток равен всей сумме документа.

    Обязательство на конец месяца — законное состояние продукта, а не пробел
    в демо-данных: партнёр обязан видеть его на экране счетов каждый день.
    """
    balances = _invoice_balances(conn)
    assert balances, "в демо нет ни одного счёта поставщика"
    unpaid = [amount for amount, paid in balances if paid == 0]
    assert any(amount > 0 for amount in unpaid), (
        f"нет счёта без единой оплаты: {balances}"
    )


def test_there_is_a_partially_paid_invoice(conn):
    """Остаток частично оплаченного счёта — число строго между нулём и суммой.

    Не «оплачено да/нет»: платёж меньше документа — законный промежуточный
    остаток, и именно так проверяется вариант развилки Q018 про частичную
    оплату.
    """
    balances = _invoice_balances(conn)
    partial = [(amount, paid) for amount, paid in balances if 0 < paid < amount]
    assert partial, f"нет частично оплаченного счёта: {balances}"


def test_the_classification_inbox_is_not_empty(conn):
    """В инбоксе классификации (T152) есть действующая строка без статьи.

    Строка без статьи получает служебный код `unclassified` и остаётся
    видимой числом — не пропадает и не превращается в ноль.
    """
    rows = conn.execute(
        """select f.amount
             from facts f join pnl_items i on i.id = f.pnl_item_id
            where i.code = 'unclassified' and f.superseded_at is null"""
    ).fetchall()
    assert rows, "в демо нет ни одной строки без статьи — инбокс пуст"
    assert any(amount > 0 for amount, in rows), (
        f"строка без статьи есть, но её сумма не положительна: {rows}"
    )


def test_a_supplier_payment_does_not_double_the_expense(conn):
    """Платёж не удваивает расход: перевод и начисление считаются раздельно.

    `pnl_lines` сама по виду не фильтрует — это делает каждый потребитель
    (`reports.export.EXPENSE_LINES` и остальные, T151). Счёт №5 и его платёж —
    единственная пара демо, лежащая в одном периоде (август): не исключи
    кто-нибудь `kind = 'transfer'`, расход посчитался бы дважды, и именно на
    этой паре ошибка была бы видна.
    """
    row = conn.execute(
        """select coalesce(sum(amount) filter (where kind <> 'transfer'), 0),
                  coalesce(sum(amount) filter (where kind = 'transfer'), 0)
             from pnl_lines
            where fact_id in (
                select id from facts
                 where dedup_key in ('manual:invoice:demo-inv-5', 'manual:payment:demo-pay-3')
            )"""
    ).fetchone()
    expense_total, transfer_total = row
    assert transfer_total == Decimal("23900.00"), (
        f"платёж не найден среди переводов в pnl_lines: {transfer_total}"
    )
    assert expense_total == Decimal("23900.00"), (
        f"расход посчитан не той суммой — платёж мог удвоить его: {expense_total}"
    )


def test_the_pnl_export_of_a_closed_month_has_three_parts(demo_db):
    """Выгрузка закрытого месяца показывает зарплату, наличные и поставщика разом.

    Зарплата видна по `kind = 'Accrual'`. Наличные и счета поставщиков оба идут
    как `Expense`, и разделяются по статье: `Food cost` в демо получает
    исключительно счёт поставщика (Metro Cash & Carry, T151) — у трат из кассы
    такой статьи нет вовсе. Её появление в июльской выгрузке и есть проверка,
    что счёт поставщика долетает до P&L, а не только до списка счетов на экране.
    """
    result = report(demo_db, """
import io, json
from datetime import date
from django.utils import translation
from core.models import Tenant
from reports.export import pnl
from reports.sheet import build_slice
from web.cash import item_title
t = Tenant.objects.get(code="demo")
with translation.override("en"):
    body, name = pnl(
        build_slice(t.id, date(2026, 7, 1)), tenant_id=t.id, period=date(2026, 7, 1),
        title="July 2026", item_title=item_title,
    )
import openpyxl
sheet = openpyxl.load_workbook(io.BytesIO(body)).active
rows = [r for r in sheet.iter_rows(values_only=True) if r and r[3]]
print(json.dumps({"kinds": sorted({r[3] for r in rows}),
                  "articles": sorted({r[0] for r in rows if r[3] == "Expense"})},
                 ensure_ascii=False))
""")
    assert "Accrual" in result["kinds"], f"в выгрузке июля нет зарплаты: {result}"
    assert "Food cost" in result["articles"], (
        f"счёт поставщика не долетел до выгрузки P&L: {result['articles']}"
    )
    assert any(article != "Food cost" for article in result["articles"]), (
        f"в выгрузке нет расходов из кассы кроме поставщика: {result['articles']}"
    )
