"""Расходы в выгрузке «Строки для P&L» рядом с зарплатными (T113).

Ради чего вообще собираются расходы: чтобы в P&L не было дыры ровно там, где
деньги тратятся мимо банка. Дыра закрывается не тем, что расход записан, а тем,
что он **доехал до файла, который бухгалтер собирает в P&L**. Пока он туда не
доехал, продукт умеет копить траты и не умеет их отдавать.

Отсюда три вещи, которые проверяются здесь и которых не проверяет ни один
соседний файл.

**1. Обе части в одном файле и в одних статьях.** Зарплата приезжает из
ведомости (`SheetSlice`), расходы — из фактов; статья у них одна и та же
сущность (`pnl_items`), и в файле они обязаны стоять строками одного формата.
Файл, где расходы лежат отдельным листом со своими колонками, для сборки P&L
бесполезен так же, как файл без расходов.

**2. Ни одна сумма не пропадает молча.** Расход без точки (ждёт разнесения) в
файле есть — с прочерком вместо точки. Пропавшая строка — это не найденная
позже недостача, и хуже неё только строка, пропавшая тихо.

**3. Зарплата не удваивается.** Зарплатная часть берётся из ведомости, и
факты с источником `payroll` в файл не попадают. Это не гипотетическая
осторожность: проводка зарплаты в факты — отложенный долг (см. журнал блока
`facts`, T107), и в тот день, когда её сделают, файл обязан не начать считать
одни и те же деньги дважды.
"""
from __future__ import annotations

import io
import json
from decimal import Decimal

import openpyxl
import pytest

from conftest import body, login_as, period_url
from payrun.sheet import Cell, assemble
from reports.export import ExpenseLine, pnl
from reports.sheet import ALL, SheetSlice
from test_cash_expense import (  # noqa: F401
    JUNE_DAY,
    entry_key,
    facts_removed,
    item,
    payload,
    tenant,
    units,
)
from test_directory import payruns_restored, sql  # noqa: F401

D = Decimal
JUNE = "2026-06-01"
NEW = "/expenses/new/"

PAYROLL_ARTICLE = "Зарплата производственного персонала"
EXPENSE_ARTICLE = "Коммунальные"


def payroll_sheet(cut: str = ALL) -> SheetSlice:
    """Ведомость из одной суммы: зарплатная половина файла."""
    cell = Cell(
        employee="ANDRIC UROS", unit="NS1", ledger="official",
        code="hours.regular", title="Отработанные", amount=Decimal("1000.00"),
        key="dev-emp-1",
    )
    return SheetSlice(sheet=assemble([cell]), cut=cut, cuts=[])


def book_of(expenses, *, cut: str = ALL) -> list[tuple]:
    """Строки готового файла. Читаются ячейки, а не байты: xlsx — это zip."""
    body, _name = pnl(
        payroll_sheet(cut), title="Июнь 2026",
        articles={"dev-emp-1": PAYROLL_ARTICLE}, taxes=[], expenses=expenses,
    )
    sheet = openpyxl.load_workbook(io.BytesIO(body)).active
    return [
        row for row in sheet.iter_rows(values_only=True)
        if any(value is not None for value in row)
    ]


def test_expense_rows_stand_next_to_payroll_rows_in_the_same_articles():
    """Обе части в одном файле, одного формата и с разными типами строки."""
    rows = book_of([
        ExpenseLine(EXPENSE_ARTICLE, "NS1", "official", "Вода", Decimal("120.50")),
    ])

    kinds = {row[3] for row in rows if row[3]}
    assert "Начисление" in kinds, rows
    assert "Расход" in kinds, f"расходов в файле нет вовсе: {rows}"

    expense = [row for row in rows if row[3] == "Расход"]
    assert len(expense) == 1, expense
    assert expense[0][0] == EXPENSE_ARTICLE, "расход попал не в свою статью"
    assert expense[0][1] == "NS1"
    assert expense[0][4] == "Вода", "не названо, за что деньги"

    accrual = [row for row in rows if row[3] == "Начисление"]
    assert accrual and accrual[0][0] == PAYROLL_ARTICLE
    # Ширина строки одна: загрузчик P&L читает файл одним разбором.
    assert len(expense[0]) == len(accrual[0])


def test_an_expense_without_a_unit_is_in_the_file_with_a_dash():
    """Нераспределённая сумма из файла не пропадает и не притворяется чужой.

    Пустая ячейка читается как «забыли заполнить», прочерк — как «точка ещё не
    решена». Разница в том, будет ли эта сумма кем-то найдена.
    """
    rows = book_of([
        ExpenseLine(EXPENSE_ARTICLE, "", "official", "Аренда офиса", Decimal("999.99")),
    ])
    waiting = [row for row in rows if row[3] == "Расход"]
    assert waiting, f"расход без точки исчез из файла: {rows}"
    assert waiting[0][1] == "—", waiting[0]


def test_the_money_reaches_the_file_to_the_kopeck():
    """Сумма в ячейке ровно та, что записана: без двоичного хвоста и локализации."""
    rows = book_of([
        ExpenseLine(EXPENSE_ARTICLE, "NS1", "official", "Вода", Decimal("80756.32")),
    ])
    amounts = [str(row[5]) for row in rows if row[3] == "Расход"]
    # Строкой, а не числом: `Decimal(...) == 80756.32` было бы зелено и на
    # двоичном хвосте, ради лечения которого написан `_money_format`.
    assert amounts == ["80756.32"], amounts


def test_expense_rows_of_one_kind_are_summed_not_repeated():
    """Одинаковые строки складываются: файл — заготовка P&L, а не журнал."""
    same = ExpenseLine(EXPENSE_ARTICLE, "NS1", "official", "Вода", Decimal("10.01"))
    rows = book_of([same, same])
    expense = [row for row in rows if row[3] == "Расход"]
    assert len(expense) == 1, expense
    assert str(expense[0][5]) == "20.02", expense[0]


# --- на живой базе: что доезжает до файла каждой роли --------------------------
#
# Выше проверен формат файла, здесь — **срез**. Он делается политиками базы, а не
# выборкой (D014), и потому проверяется теми же тремя ролями, которыми
# проверяется экран: файл уходит из продукта и живёт своей жизнью, поэтому лишняя
# строка в нём дороже лишней строки на экране (D023).

def june_export_url(sql, tenant) -> str:  # noqa: F811
    """Адрес выгрузки июня. Номер периода из базы, а не со страницы списка.

    Со списка приезжает «первый попавшийся» период, и тест зависел бы от того,
    какой месяц продукт показывает сегодня, — то есть падал бы в первый день
    следующего месяца, ничего не найдя в файле.
    """
    period_id = sql.execute(
        "select id from periods where tenant_id = %s and period = %s", (tenant, JUNE)
    ).fetchone()[0]
    return f"/periods/{period_id}/export/pnl/"


def spend(client, item_id, units_map, **extra) -> None:
    """Внести расход июньской датой — тем же путём, что человек."""
    form = payload(item_id, units_map, entry_key=entry_key(), date=JUNE_DAY, **extra)
    answer = client.post(NEW, form)
    assert answer.status_code == 302, body(answer)


def money_of(rows) -> list[Decimal]:
    """Суммы строк числами: сверяются деньги, а не их запись в ячейке."""
    return [Decimal(str(row[5])) for row in rows]


def expense_rows(client, url) -> list[tuple]:
    response = client.get(url)
    assert response.status_code == 200, response.status_code
    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
    return [
        row for row in sheet.iter_rows(values_only=True)
        if row and row[3] == "Расход"
    ]


@pytest.fixture
def three_expenses(client, sql, item, units, facts_removed):  # noqa: F811
    """Три июньских расхода на трёх точках, внесённые бухгалтером с экрана."""
    login_as(client, "accountant")
    for code, amount in (("NS1", "100.00"), ("NS2", "200.00"), ("BG1", "300.00")):
        spend(client, item, units, unit=units[code], amount=amount, note=f"вода {code}")
    client.post("/logout/")


def test_the_accountant_gets_every_unit_and_the_manager_only_his_own(
    client, sql, tenant, three_expenses,  # noqa: F811
):
    """Файл — это срез смотрящего, и срез делает база.

    Сломайте `unit_visibility` на `facts` — в файле управляющего появятся чужие
    точки и чужие суммы, и он соберёт P&L по данным, которых ему не показывали.
    """
    url = june_export_url(sql, tenant)

    login_as(client, "accountant")
    whole = expense_rows(client, url)
    assert sorted(money_of(whole)) == [D("100"), D("200"), D("300")], whole
    client.post("/logout/")

    login_as(client, "manager")
    mine = expense_rows(client, url)
    assert [row[1] for row in mine] == ["NS1"], mine
    assert money_of(mine) == [D("100")], mine


def test_the_ledger_cut_narrows_the_expenses_too(
    client, web_env, sql, tenant, item, units, facts_removed, payruns_restored,  # noqa: F811
):
    """Разрез по регистру сужает и расходы, а не только зарплату.

    Иначе файл разреза показывал бы зарплату одного регистра и расходы всех —
    и итог по нему не сошёлся бы ни с чем.

    Месяц здесь считается намеренно: словарь разрезов собирается из регистров
    **ведомости** (`reports.sheet.slice_cells`), и у периода без расчёта любой
    разрез схлопывается во «все видимые». Это поведение старше расходов и здесь
    не меняется — но знать о нём тест обязан, иначе он проверял бы отсутствие
    разреза, а не сужение (issue #108).
    """
    login_as(client, "director")
    assert client.post(
        period_url(client) + "calculate/", {"inline": "1"}, follow=True
    ).status_code == 200
    spend(client, item, units, unit=units["NS1"], amount="50.00", ledger="official")
    spend(client, item, units, unit=units["NS1"], amount="70.00", ledger="internal")

    url = june_export_url(sql, tenant)
    whole = money_of(expense_rows(client, url))
    official = money_of(expense_rows(client, url + "?ledger=official"))

    assert sorted(whole) == [D("50"), D("70")], whole
    assert official == [D("50")], official


def test_payroll_facts_do_not_land_in_the_file_a_second_time(
    client, sql, tenant, item, units, facts_removed,  # noqa: F811
):
    """Зарплата приезжает из ведомости; факт с источником `payroll` — не приезжает.

    Проводки зарплаты в факты сегодня нет (отложена, см. журнал блока), но файл
    обязан быть готов к её появлению: иначе в день проводки одни и те же деньги
    молча удвоятся, и увидят это по расхождению P&L с ведомостью.
    """
    login_as(client, "director")
    spend(client, item, units, unit=units["NS1"], amount="100.00")

    line = sql.execute("select id from pnl_items where code = 'labour_cost'").fetchone()[0]
    # Через `upsert_fact`, а не прямым `insert`: проводка зарплаты, когда её
    # сделают, пойдёт этим же путём, и тест обязан класть то же, что она.
    sql.execute(
        "select fact_id from upsert_fact(%s::jsonb)",
        (json.dumps({
            "tenant_id": tenant, "period": JUNE, "doc_date": JUNE_DAY,
            "unit_id": units["NS1"], "pnl_item_id": str(line), "ledger": "official",
            "amount": "777.77", "title": "Зарплата, начисления", "channel": "bank",
            "source": "payroll", "dedup_key": "payroll:test", "allocation": "direct",
        }),),
    )
    try:
        rows = expense_rows(client, june_export_url(sql, tenant))
    finally:
        sql.execute("delete from facts where dedup_key = 'payroll:test'")

    assert money_of(rows) == [D("100")], (
        f"зарплатный факт приехал в файл вторым разом: {rows}"
    )
