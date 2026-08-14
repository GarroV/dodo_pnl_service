"""Статья P&L в выгрузке ищется тем же ключом, каким собран справочник (issue #95).

Что было. `pnl()` брал статью так:

    article = articles.get(row.employee, _("Без статьи"))

`articles` собран по ключу сотрудника (`employees.external_id`), а `row.employee` —
**отображаемое** имя строки ведомости, «ФАМИЛИЯ ИМЯ». Совпасть они не могли в
принципе, поэтому у всех начислений в файле стояло «Без статьи»: на стенде 26
строк из 32. Налоговые строки не задеты — там спрашивали правильным ключом, и
именно поэтому дефект выглядел как «статья есть, но не у всех», а не как поломка.

Почему это важнее, чем кажется. P&L — цель всего продукта; выгрузка, где все
начисления лежат под одной служебной статьёй, для сборки P&L бесполезна и молчит
об этом: файл выглядит нормальным.

Почему тест такой. Он идёт от **строки ведомости**, а не от функции разбора: сам
дефект был в том, что два места спрашивали одно и то же разными ключами, и
проверка одного из них по отдельности осталась бы зелёной.
"""
from __future__ import annotations

from decimal import Decimal

from payrun.sheet import Cell, assemble
from reports.export import pnl
from reports.sheet import ALL, SheetSlice

KEY = "dev-emp-1"
SHOWN = "ANDRIC UROS"
ARTICLE = "Зарплата производственного персонала"


def sheet_of_one() -> SheetSlice:
    """Ведомость из одной суммы: ключ сотрудника и его отображаемое имя РАЗНЫЕ."""
    cell = Cell(
        employee=SHOWN, unit="NS1", ledger="official",
        code="hours.regular", title="Отработанные", amount=Decimal("1000.00"),
        key=KEY,
    )
    return SheetSlice(sheet=assemble([cell]), cut=ALL, cuts=[])


def values(articles: dict[str, str]) -> set[str]:
    """Что в файле НА САМОМ ДЕЛЕ: ячейки, а не байты.

    Искать текст в байтах xlsx нельзя: это zip, и строка внутри сжата. Первая
    версия этой проверки так и делала — и её вторая половина зеленела на пустом
    месте, потому что «текста нет в архиве» верно всегда.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    # `expenses=[]` рядом с `taxes=[]`, и это не украшение: без них выгрузка
    # пошла бы в базу за расходами (T113), а проверка здесь про другое — про
    # то, каким ключом ищется статья. Тест без базы ловит это быстрее и точнее.
    body, _name = pnl(
        sheet_of_one(), title="Июнь 2026", articles=articles, taxes=[], expenses=[],
    )
    sheet = load_workbook(BytesIO(body)).active
    return {str(cell) for row in sheet.iter_rows(values_only=True) for cell in row if cell}


def test_the_article_reaches_the_line_of_an_accrual():
    """Главная проверка: статья доезжает до строки начисления, а не теряется."""
    assert ARTICLE in values({KEY: ARTICLE}), (
        "статья группы не доехала до строки начисления — в файле осталось «Без статьи»"
    )


def test_the_display_name_is_not_used_as_a_key():
    """Порча наоборот: справочник по отображаемому имени НЕ должен срабатывать.

    Без этой проверки починку можно было бы «сделать» вторым справочником по
    имени — и однофамильцы получили бы чужую статью молча.
    """
    assert ARTICLE not in values({SHOWN: ARTICLE}), (
        "статья нашлась по отображаемому имени — значит ключ по-прежнему не тот"
    )
