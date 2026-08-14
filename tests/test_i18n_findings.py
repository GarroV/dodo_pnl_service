"""Находки разбора таблицы партнёра — на языке страницы (T140, issue #96).

Раздел «Не разобрано в файле» на странице сверки и отчёт импорта табеля печатали
находки **по-русски на любом языке**: тексты собирались f-строками мимо каталога,
а место в файле склеивалось как `f"{лист}, строка {номер}"` — слово «строка» тоже
жёстко русское.

Почему это не поймали остальные проверки локализации. `tests/test_i18n.py` ищет
кириллицу на **отрисованных** экранах, а находки появляются только после
загрузки файла, который разбирается не до конца; обычные смоуки грузят
корректный файл, где находок нет вовсе. То есть дыра была ровно в том месте, где
её никто не искал: демо целиком англоязычное (D035), и стоило смотрящему
загрузить чужой файл — он получал русский абзац.

Проверяется здесь то же, что и везде в локализации: не «перевод существует», а
**отсутствие непереведённого**. Язык исходника русский, поэтому кириллица в
находке на английском языке и есть точный признак строки мимо каталога.
"""
from __future__ import annotations

import re
import shutil

import pytest
from django.utils.translation import override

from conftest import PLATA_SAMPLE

CYRILLIC = re.compile(r"[а-яёА-ЯЁ]")

SHEET = "NS 1 Bulevar "
OTHER_SHEET = "NS 2 Dunavska "


@pytest.fixture
def broken(tmp_path):
    """Копия эталона со всеми видами беды сразу — по одной на каждый вид находки."""
    import openpyxl

    path = tmp_path / "broken.xlsx"
    shutil.copy(PLATA_SAMPLE, path)
    book = openpyxl.load_workbook(path)

    # Данные в файле — латиницей намеренно: тогда любая кириллица в находке
    # означает текст продукта мимо каталога, и исключений в проверке не нужно.
    book.create_sheet("Zabeleške knjigovodje")       # лист не из формата
    del book[OTHER_SHEET]                            # лист формата пропал

    sheet = book[SHEET]
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(1, col).value or "").strip().upper().startswith("SATI RADA"):
            sheet.cell(1, col).value = "SATI (novo)"  # колонка переименована
            break
    else:  # pragma: no cover — эталон обязан содержать эту колонку
        pytest.fail("в эталоне нет колонки часов")

    sheet.cell(2, 1).value = None                     # строка без номера
    sheet.cell(3, 2).value = None                     # строка без имени
    sheet.cell(3, 3).value = None

    other = book["BG1 pun obracun "]
    for col in range(1, other.max_column + 1):
        if str(other.cell(1, col).value or "").strip().upper() == "KOEFICIJENT":
            other.cell(4, col).value = "osam"        # не число
            break
    else:  # pragma: no cover
        pytest.fail("в эталоне нет колонки коэффициента")

    book.save(path)
    return path


def findings(path):
    from payroll.importers import read_plata_file

    return read_plata_file(path).findings


def test_the_broken_file_gives_every_kind_of_finding(broken):
    """Сначала убеждаемся, что проверять есть что: иначе тест языка пуст."""
    kinds = {finding.kind for finding in findings(broken)}
    assert kinds == {"sheet", "column", "row", "value"}, kinds


def test_findings_speak_english_on_an_english_page(broken):
    """Ни одной кириллической буквы — ни в тексте находки, ни в месте в файле.

    Данные самого файла (имена листов, значения ячеек) в этой копии латиницей,
    поэтому исключать из проверки нечего: всё, что осталось русским, — текст
    продукта, не доехавший до каталога.
    """
    with override("en"):
        russian = [
            (finding.kind, finding.where, finding.text)
            for finding in findings(broken)
            if CYRILLIC.search(finding.text) or CYRILLIC.search(finding.where)
        ]
    assert russian == []


def test_the_place_in_the_file_is_one_translated_string(broken):
    """«Лист, строка N» — одна строка перевода с подстановками, а не склейка.

    Склеенное место переводится наполовину: имя листа приезжает из файла, а слово
    «строка» остаётся русским навсегда. У порядка слов в языках свои правила —
    поэтому строка перевода целиком, как уже сделано в `reports/own_export.py`.
    """
    with override("en"):
        places = [f.where for f in findings(broken) if f.kind in ("row", "value")]
    assert places, "не на чем проверять — файл не дал ни строчных, ни значенческих находок"
    for place in places:
        assert not CYRILLIC.search(place), place
        assert re.search(r"\d+$", place), place


def test_findings_stay_russian_on_the_russian_page(broken):
    """Обратная сторона: русский текст не должен пропасть вместе с починкой."""
    with override("ru"):
        texts = " ".join(f.text for f in findings(broken))
    assert "не загружен" in texts
    assert "не найдена" in texts
