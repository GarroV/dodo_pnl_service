"""Своя выгрузка читается своей же сверкой (T119).

**Что было.** Бухгалтер скачивал «Вид бухгалтера» и тут же загружал этот файл
на сверку. Ответ: «Сошлось до копейки 0», восемь строк «лист не входит в формат
PLATA» и поимённый список всех 35 человек с формулировкой «в загруженной
таблице такой строки нет» — при том что их имена в файле есть, человек их
видит.

Причина: сверка знала один формат — таблицу партнёра (`plata_xlsx`), у которой
свои имена листов, заголовки в первой строке и имя с фамилией в двух колонках.
Наша выгрузка устроена иначе, и разбор честно не находил в ней ничего.

**Почему это важно.** Спека: «бухгалтер загружает свою же таблицу за прошлый
месяц, система считает параллельно и показывает построчное сравнение — это то,
что превращает недоверие в доверие» (D024). Первое, что делает человек,
получивший файл, — возвращает его обратно. Продукт отвечал списком фамилий и
словами «такой строки нет», и это выглядит как поломка расчёта.

**Как проверяется.** Тем же путём, каким ходит человек: скачать выгрузку со
страницы периода и загрузить её на страницу сверки. Проверка «функция чтения
понимает функцию записи» этого не показала бы — она обходит и маршрут, и
разрез, и сопоставление имён, то есть ровно то место, где всё и ломалось.
"""
from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from conftest import PLATA_SAMPLE, body, login_as, period_url, wipe_payruns
from test_reports_reconcile_db import section_rows, summary

D = Decimal


@pytest.fixture
def calculated_june(client, web_env):
    """Посчитанный июнь на данных сида — общий материал для проверок ниже.

    Своя копия в каждом файле проверок намеренно: фикстура, импортированная из
    соседнего модуля, тянет за собой его порядок и его материал, а стоит она
    пять строк.
    """
    wipe_payruns(web_env)
    login_as(client, "director")
    response = client.post(period_url(client) + "calculate/", follow=True)
    assert response.status_code == 200
    return None


def export_bytes(client, user: str, kind: str = "partner") -> bytes:
    """Скачать выгрузку так, как её скачивает человек, — со страницы периода."""
    login_as(client, user)
    response = client.get(period_url(client) + f"export/{kind}/")
    assert response.status_code == 200, f"{user}: выгрузка ответила {response.status_code}"
    return response.content


def reconcile_with(client, user: str, data: bytes) -> str:
    """Загрузить книгу на страницу сверки — тем же полем, что человек."""
    login_as(client, user)
    upload = SimpleUploadedFile("vid-buhgaltera.xlsx", data)
    response = client.post(period_url(client) + "reconcile/", {"table": upload}, follow=True)
    assert response.status_code == 200, f"{user}: сверка ответила {response.status_code}"
    return body(response)


def edited(data: bytes, name: str, shift: Decimal) -> bytes:
    """Подвинуть сумму одного человека в книге — как это сделал бы Excel.

    Ровно то, ради чего сверка своей же выгрузки и нужна: человек выгружает наш
    расчёт, правит в Excel числа на свои и приносит обратно.
    """
    book = openpyxl.load_workbook(io.BytesIO(data))
    changed = 0
    for ws in book.worksheets:
        for row in ws.iter_rows():
            # Правится ПЕРВАЯ строка человека: строк у него столько, сколько
            # регистров, а бухгалтер в Excel правит одну ячейку — и итог по
            # человеку обязан поехать ровно на неё.
            if changed or str(row[1].value or "").strip() != name:
                continue
            cell = row[-1]
            cell.value = D(str(cell.value)) + shift
            changed += 1
    assert changed == 1, f"строк «{name}» в книге не нашлось"
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


# --- приёмка: файл сходится сам с собой ---------------------------------------


def test_our_own_view_is_read_by_our_own_reconciliation(client, calculated_june):
    """Скачал «Вид бухгалтера» — загрузил обратно — сошлось построчно."""
    html = reconcile_with(client, "accountant", export_bytes(client, "accountant"))
    counts = summary(html)

    assert counts["Сошлось до копейки"] == 35, f"сводка сверки: {counts}"
    assert counts["Разошлось"] == 0, f"сводка сверки: {counts}"
    assert counts["Разошлось на копейки (округление)"] == 0, f"сводка сверки: {counts}"
    assert counts["Сверены только входы — деньги не сравнивались"] == 0, counts
    assert counts["Есть в таблице, нет в вашей части расчёта"] == 0, counts
    assert counts["Есть в расчёте, нет в таблице"] == 0, counts
    assert "не входит в формат PLATA" not in html, (
        "своя выгрузка по-прежнему читается как чужая таблица"
    )
    assert "Не разобрано в файле" not in html, "в своей же выгрузке что-то не разобрано"


def test_a_number_changed_in_excel_comes_back_as_a_difference(client, calculated_june):
    """Тот самый сценарий переезда: правку в Excel сверка показывает построчно."""
    original = export_bytes(client, "accountant")
    html = reconcile_with(client, "accountant", edited(original, "ANDRIC UROS", D("-100.00")))
    counts = summary(html)

    assert counts["Разошлось"] == 1, f"правка на 100 динаров не найдена: {counts}"
    assert counts["Сошлось до копейки"] == 34, f"сводка сверки: {counts}"
    off = section_rows(html, "Разошлось")
    assert any("ANDRIC UROS" in " ".join(row) for row in off), off
    assert any("100,00" in " ".join(row) for row in off), (
        f"разница показана не тем числом: {off}"
    )


def test_the_partner_table_is_still_read_as_before(client, calculated_june):
    """Второй формат не отменяет первого: таблица партнёра читается как читалась."""
    login_as(client, "accountant")
    with PLATA_SAMPLE.open("rb") as handle:
        response = client.post(
            period_url(client) + "reconcile/", {"table": handle}, follow=True
        )
    counts = summary(body(response))
    assert counts["Сошлось до копейки"] == 32, f"сводка сверки: {counts}"


# --- D023 на новой поверхности ------------------------------------------------


def test_a_role_without_the_whole_run_is_told_nothing_about_it(client, calculated_june):
    """Управляющий не получает ни имён, ни чисел из чужой части расчёта (T100).

    Новый разбор — новая поверхность, и правило то же: список «есть в расчёте,
    нет в таблице» строится только для роли, которой отдан весь расчёт. У
    управляющего его нет, поэтому в сводке стоит слово, а не ноль: ноль —
    такое же утверждение о расчёте, и проверить его роль не может.
    """
    html = reconcile_with(client, "manager", export_bytes(client, "manager"))
    counts = summary(html)

    assert counts["Есть в расчёте, нет в таблице"] == "не проверялось", counts
    assert "Внутренний" not in html and "internal" not in html
    for name in ("Курир Ана", "Курир Марко", "KURIR ANA", "KURIR MARKO"):
        assert name not in html, f"на сверке управляющего назван «{name}»"


def test_a_row_without_compared_inputs_does_not_claim_they_agree(client, calculated_june):
    """«Входы сошлись» про файл, в котором входов нет вовсе, — неправда.

    В нашей выгрузке нет ни часов, ни ставки: там только деньги. Роль, которой
    итоги не отданы, сравнить не может ничего — и сказать об этом обязана
    прямо, а не подписью «входы сошлись с таблицей» под каждой строкой.
    """
    html = reconcile_with(client, "manager", export_bytes(client, "manager"))
    assert "Входы сошлись с таблицей." not in html, (
        "сверка объявила сошедшимися входы, которых в файле нет"
    )
    # И обратная половина той же неправды: «в таблице 0 часов» про файл, где
    # колонок с часами нет вовсе. Ноль здесь — не значение, а выдумка.
    assert "Часы ·" not in html, "сверка назвала расхождением часы, которых в файле нет"
    assert "Сверено то, что от регистра не зависит" not in html, (
        "раздел обещает сверку входов, которых в файле нет"
    )


# --- разбор формата -----------------------------------------------------------


def test_the_partner_table_is_not_mistaken_for_our_view():
    """Опознание формата не должно срабатывать на чужой книге."""
    from reports.own_export import looks_like_own_export

    book = openpyxl.load_workbook(PLATA_SAMPLE, data_only=True)
    assert not looks_like_own_export(book)


def test_our_view_is_recognised_and_read(client, calculated_june):
    """Книга опознана, строки прочитаны, суммы — те же, что в ней стоят."""
    from reports.own_export import looks_like_own_export, read_own_export

    data = export_bytes(client, "accountant")
    book = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    assert looks_like_own_export(book)

    parsed = read_own_export(book)
    assert not parsed.findings, f"в своей же выгрузке находки разбора: {parsed.findings}"
    assert len(parsed.rows) == 60, (
        f"строк прочитано {len(parsed.rows)}, а в ведомости их 60"
    )
    assert sum((row.total for row in parsed.rows), D(0)) == D("1951806.13")


def test_a_sheet_that_is_not_ours_is_named_and_not_swallowed(client, calculated_june):
    """Лист, который разбор не понял, назван — иначе часть файла пропадёт молча."""
    from reports.own_export import read_own_export

    data = export_bytes(client, "accountant")
    book = openpyxl.load_workbook(io.BytesIO(data))
    book.create_sheet("Заметки")["A1"] = "что-то своё"
    out = io.BytesIO()
    book.save(out)

    parsed = read_own_export(openpyxl.load_workbook(io.BytesIO(out.getvalue()), data_only=True))
    assert [f.where for f in parsed.findings] == ["Заметки"], parsed.findings


def test_two_people_with_the_same_name_are_not_merged_in_silence():
    """Однофамильцы с одинаковым именем сводятся в одну строку — и об этом сказано.

    Имени в нашей выгрузке хватает для сопоставления ровно до тех пор, пока оно
    различает людей. Сквозного ключа в файле нет намеренно (D007): в
    продуктовой постановке это JMBG, и файлу, который уходит из продукта, он не
    нужен. Значит остаётся честно сказать, что двоих различить не удалось.
    """
    from reports.reconcile import RunLine, run_by_name

    run = {
        "1": RunLine(key="1", name="PETROVIC MARKO", totals={"net": D(1)}, hours={}),
        "2": RunLine(key="2", name="PETROVIC MARKO", totals={"net": D(2)}, hours={}),
    }
    by_name, findings = run_by_name(run)

    assert "PETROVIC MARKO" not in by_name, "однофамильцы сведены молча"
    assert len(findings) == 1 and findings[0].where == "PETROVIC MARKO"


# --- подпись причины расхождения (issue #75) ----------------------------------


def test_the_base_rate_is_not_called_an_hourly_rate():
    """У сдельной группы `employees.base_rate` — цена доставки, а не часа."""
    from web.reports_views import CAUSE_TITLES

    assert CAUSE_TITLES["rate"] == "Базовая ставка"
