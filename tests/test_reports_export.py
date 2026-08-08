"""Три выгрузки периода в xlsx (T032).

Что здесь проверяется и почему именно так.

**Файл читается обратно, а не «скачался».** Каждая проверка открывает
получившуюся книгу через openpyxl и складывает числа из ячеек — то же, что
сделает человек в Excel. Проверка «ответ вернулся длиной 9 килобайт» не
означает ничего: файл может быть пустым, кривым или чужим.

**Выгрузка не содержит того, чего нет на экране.** Файл уходит из продукта и
живёт своей жизнью, поэтому регистр, которого роль не видит, не должен
попадать в него ни строкой, ни названием, ни вкладом в итог (D023). Проверяется
не по внутренним объектам, а по **тексту книги целиком**.

**Итог файла равен сумме его строк.** Тот же довод, что у ведомости: итог,
больший суммы показанного, выдаёт скрытое вычитанием.
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from payrun.sheet import Cell
from reports import export as exports
from reports.export import TaxLine
from reports.sheet import slice_cells

D = Decimal


def cell(employee, ledger, code, title, amount, unit="NS1", **kw) -> Cell:
    return Cell(
        employee=employee, unit=unit, ledger=ledger, code=code, title=title,
        amount=D(amount), key=employee, **kw,
    )


MATERIAL = [
    cell("Иванов Иван", "official", "hours.regular", "Отработанные", "60000.00"),
    cell("Иванов Иван", "official", "meal_and_vacation_bonus", "Топли оброк", "1500.00"),
    cell("Петров Пётр", "supplementary", "hours.regular", "Отработанные", "40000.00", unit="BG1"),
    cell("Петров Пётр", "official", "meal_and_vacation_bonus", "Топли оброк",
         "1500.00", unit="BG1"),
    cell("Сидоров Сидор", "internal", "hours.regular", "Отработанные", "25000.00"),
]

LEDGER_TITLES = {
    "official": "Официальный",
    "supplementary": "Дополнительный",
    "internal": "Внутренний",
}


def titles(code: str) -> str:
    return LEDGER_TITLES.get(code, code)


def build(kind: str, cells=None, cut: str = "", **kw):
    view = slice_cells(list(cells if cells is not None else MATERIAL), cut)
    if kind == "pnl":
        # Статьи и налоги приезжают из базы; здесь их подставляют явно, чтобы
        # проверять сборку файла, а не выборку. Путь из базы проверяется
        # отдельно, на живом стенде и ролью `app_user`.
        kw.setdefault("articles", {})
        kw.setdefault("taxes", [])
    body, name = getattr(exports, kind)(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=titles, **kw,
    )
    return openpyxl.load_workbook(io.BytesIO(body)), name


def numbers(ws) -> list[Decimal]:
    return [
        Decimal(str(value))
        for row in ws.iter_rows(values_only=True)
        for value in row
        if isinstance(value, (int, float, Decimal))
    ]


def text_of(book) -> str:
    """Весь текст книги — так утечку названия регистра видно в любом углу файла."""
    return "\n".join(
        str(value)
        for ws in book.worksheets
        for row in ws.iter_rows(values_only=True)
        for value in row
        if value is not None
    )


def column(ws, header: str) -> list:
    """Значения колонки по её заголовку — как её находит глазами человек."""
    for row in ws.iter_rows(values_only=True):
        if header in [str(v) for v in row if v is not None]:
            index = [str(v) if v is not None else None for v in row].index(header)
            break
    else:
        raise AssertionError(f"в листе «{ws.title}» нет колонки «{header}»")

    out, started = [], False
    for row in ws.iter_rows(values_only=True):
        if started and row[index] is not None:
            out.append(row[index])
        if header in [str(v) for v in row if v is not None]:
            started = True
    return out


# --- ведомость к выплате -----------------------------------------------------


def test_payout_opens_and_holds_every_visible_row():
    book, _ = build("payout")
    ws = book.active

    people = column(ws, "Сотрудник")
    assert len(people) == 4, f"строк в файле не четыре: {people}"
    assert "Иванов Иван" in people


def test_payout_total_equals_the_sum_of_its_own_rows():
    """Складываем колонку «Итого» файла и сравниваем с подвалом файла."""
    book, _ = build("payout")
    ws = book.active

    rows = [Decimal(str(v)) for v in column(ws, "Итого")]
    assert len(rows) == 5, "в колонке «Итого» должны быть четыре строки и подвал"
    *lines, footer = rows
    assert sum(lines, D(0)) == footer
    assert footer == D("128000.00")


def test_payout_carries_the_same_total_as_the_screen():
    view = slice_cells(list(MATERIAL))
    book, _ = build("payout")

    *_, footer = [Decimal(str(v)) for v in column(book.active, "Итого")]
    assert footer == view.sheet.total


def test_payout_of_a_cut_holds_only_that_cut():
    book, name = build("payout", cut="official")

    assert numbers(book.active) and sum(
        Decimal(str(v)) for v in column(book.active, "Итого")
    ) == D("63000.00") * 2, "итог разреза не сходится с суммой его строк"
    assert "official" in name, "имя файла не называет разрез"


@pytest.mark.parametrize("kind", ["payout", "pnl", "partner"])
def test_no_export_names_a_ledger_the_role_cannot_see(kind):
    """Главная проверка T032: файл не знает о чужих регистрах (D023).

    Роли видны две трети данных; в файле не должно быть ни строк третьего
    регистра, ни его названия, ни его вклада в итог.
    """
    visible = [c for c in MATERIAL if c.ledger != "internal"]
    book, _ = build(kind, cells=visible)

    assert "Внутренний" not in text_of(book)
    assert "internal" not in text_of(book)
    assert "Сидоров Сидор" not in text_of(book)
    assert D("25000.00") not in numbers(book.active)


def test_export_of_a_single_ledger_role_never_prints_the_word_cut():
    """У роли с одним регистром разреза нет, и файл о нём не заикается."""
    official = [c for c in MATERIAL if c.ledger == "official"]
    book, name = build("payout", cells=official)

    assert "Дополнительный" not in text_of(book)
    assert "Внутренний" not in text_of(book)
    assert name.endswith(".xlsx") and "official" not in name


def test_payout_marks_a_retro_row_with_its_source_month():
    """Ретро-строка обязана объяснить себя: иначе это сумма из ниоткуда."""
    material = MATERIAL + [
        cell("Иванов Иван", "official", "hours.regular", "Отработанные", "500.00",
             retro_source=date(2026, 5, 1)),
    ]
    book, _ = build("payout", cells=material)

    assert "05.2026" in text_of(book)


# --- строки для P&L ----------------------------------------------------------


def test_pnl_splits_accruals_from_taxes():
    book, _ = build(
        "pnl",
        articles={"Иванов Иван": "LC / КС", "Петров Пётр": "LC / КС",
                  "Сидоров Сидор": "LC / DC"},
        taxes=[TaxLine("LC / КС", "NS1", D("1000.00"), D("2000.00"))],
    )
    ws = book.active
    kinds = {str(v) for v in column(ws, "Тип строки")}

    assert "Начисление" in kinds
    assert "Налог" in kinds and "Взносы" in kinds


def test_pnl_accruals_equal_the_visible_slice():
    """Сумма начислений P&L обязана сойтись с итогом ведомости того же среза."""
    view = slice_cells(list(MATERIAL))
    book, _ = build("pnl", articles={}, taxes=[])
    ws = book.active

    accruals = [
        Decimal(str(amount))
        for kind, amount in zip(column(ws, "Тип строки"), column(ws, "Сумма"), strict=True)
        if str(kind) == "Начисление"
    ]
    assert sum(accruals, D(0)) == view.sheet.total


def test_pnl_of_a_cut_carries_no_taxes():
    """Налог считается по строке ведомости целиком — в разрезе регистра его нет."""
    book, _ = build(
        "pnl", cut="official",
        articles={}, taxes=[TaxLine("LC / КС", "NS1", D("1000.00"), D("2000.00"))],
    )
    kinds = {str(v) for v in column(book.active, "Тип строки")}

    assert kinds == {"Начисление"}, "в разрезе одного регистра налогов быть не может"


def test_pnl_names_the_article_of_every_line():
    book, _ = build(
        "pnl", articles={"Иванов Иван": "LC / GMC"}, taxes=[],
    )
    articles = {str(v) for v in column(book.active, "Статья P&L")}

    assert "LC / GMC" in articles
    assert "Без статьи" in articles, "человек без статьи обязан быть виден, а не пропасть"


# --- вид, привычный бухгалтеру -----------------------------------------------


def test_partner_puts_every_unit_on_its_own_sheet():
    book, _ = build("partner")

    assert {ws.title for ws in book.worksheets} == {"NS1", "BG1"}


def test_partner_keeps_the_same_money_as_the_screen():
    view = slice_cells(list(MATERIAL))
    book, _ = build("partner")

    total = sum(
        (Decimal(str(v)) for ws in book.worksheets for v in column(ws, "UKUPNO ZA ISPLATU")),
        D(0),
    )
    # На каждом листе своя строка подвала, поэтому итог в файле удваивается:
    # сумма строк плюс сумма подвалов. Проверяем ровно это.
    assert total == view.sheet.total * 2


def test_partner_uses_the_headers_of_the_partner_table():
    book, _ = build("partner")
    text = text_of(book)

    assert "TOPLI OBROK I REGRES" in text
    assert "UKUPNO ZA ISPLATU" in text


def test_partner_never_calls_money_hours():
    """У партнёра «SATI RADA» — это часы. Деньги под этим заголовком были бы ложью."""
    book, _ = build("partner")
    text = text_of(book)

    assert "SATI RADA" not in text
    assert "Отработанные" in text, "начисление за часы обязано остаться названным"
