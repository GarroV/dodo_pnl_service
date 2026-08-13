"""Деньги в выгрузке xlsx не должны хранить двоичный хвост float (часть T103).

**Дефект и почему смотрим в сам файл.** `openpyxl` сериализует число через
`"%.16g" % value` (`openpyxl.compat.safe_string`). Шестнадцать значащих цифр —
на одну больше, чем нужно для короткого представления double, поэтому в XML
книги может лежать `80756.32000000001` вместо `80756.32`, хотя `load_workbook`
вернёт из такой ячейки то же самое число `80756.32` — оба варианта после
чтения `float()`-ом дают один и тот же double, и проверка через
`load_workbook` была бы зелёной и на неисправленном коде. Поэтому здесь
распаковывается сам xlsx (`zipfile`) и читается XML листа буквально: денежных
`<v>` с более чем двумя знаками после точки быть не должно.

**Формат ячейки и то, что число — число, а не текст, проверяем отдельно.**
Фикс подменяет значение ячейки строкой на уровне сериализации, и легко
случайно оставить ячейку текстовой (тогда Excel не просуммирует её и формат
`#,##0.00` не подсветит десятичные) — обе проверки на этот случай.
"""
from __future__ import annotations

import io
import re
import zipfile
from datetime import date
from decimal import Decimal

import openpyxl

from payrun.sheet import Cell
from reports import export as exports
from reports.sheet import slice_cells

D = Decimal

# Ровно то число из дефекта: под `%.16g` даёт двоичный хвост в 11-м знаке.
AMOUNT = D("80756.32")

MATERIAL = [
    Cell(
        employee="Иванов Иван", unit="NS1", ledger="official",
        code="hours.regular", title="Отработанные", amount=AMOUNT,
        key="Иванов Иван",
    ),
]

# Своё число, отдельное от AMOUNT: у него четыре знака после запятой, и оно не
# должно попадать под старую проверку хвоста (та ищет любое число с тремя и
# более знаками после точки, а тут их ровно столько по замыслу самой величины,
# не из-за двоичного хвоста).
PRECISE_AMOUNT = D("1234.5678")

PRECISE_MATERIAL = [
    Cell(
        employee="Петров Пётр", unit="NS1", ledger="official",
        code="hours.regular", title="Отработанные", amount=PRECISE_AMOUNT,
        key="Петров Пётр",
    ),
]

# Денежные `<v>` с двоичным хвостом выглядят так: точка, две и более цифр,
# затем ещё цифры без пары нулей на конце (11+ значащих цифр после точки —
# ровно то, что оставляет "%.16g" после десятичной точки для чисел такого
# порядка). Ищем любое число с более чем двумя знаками после точки.
NUMBER_WITH_TOO_MANY_DECIMALS = re.compile(r"<v>-?\d+\.\d{3,}</v>")


def _sheet_xml(body: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(body)) as archive:
        names = [n for n in archive.namelist() if n.startswith("xl/worksheets/sheet")]
        return "\n".join(archive.read(name).decode("utf-8") for name in names)


def test_payout_xlsx_has_no_binary_tail_in_money_values():
    """Смотрим в сам XML: без этого тест был бы зелёным и на баге (см. докстринг)."""
    view = slice_cells(list(MATERIAL))
    body, _ = exports.payout(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=str,
    )

    xml = _sheet_xml(body)
    assert "80756.32000000001" not in xml
    match = NUMBER_WITH_TOO_MANY_DECIMALS.search(xml)
    assert match is None, f"в XML нашлось число с двоичным хвостом: {match.group()}"


def test_pnl_xlsx_has_no_binary_tail_in_money_values():
    """Тот же фикс, тот же дефект — но через `pnl`, а не `payout`."""
    view = slice_cells(list(MATERIAL))
    body, _ = exports.pnl(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=str, articles={}, taxes=[],
    )

    xml = _sheet_xml(body)
    assert "80756.32000000001" not in xml
    match = NUMBER_WITH_TOO_MANY_DECIMALS.search(xml)
    assert match is None, f"в XML нашлось число с двоичным хвостом: {match.group()}"


def test_partner_xlsx_has_no_binary_tail_in_money_values():
    """Тот же фикс, тот же дефект — но через `partner`, а не `payout`."""
    view = slice_cells(list(MATERIAL))
    body, _ = exports.partner(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=str,
    )

    xml = _sheet_xml(body)
    assert "80756.32000000001" not in xml
    match = NUMBER_WITH_TOO_MANY_DECIMALS.search(xml)
    assert match is None, f"в XML нашлось число с двоичным хвостом: {match.group()}"


def test_payout_xlsx_keeps_amount_with_more_than_two_decimals_exact():
    """Округление в выгрузке — молчаливый дефект: сумма строк файла разойдётся
    с его же итогом, если значение показано не тем, что было посчитано (T103).
    """
    view = slice_cells(list(PRECISE_MATERIAL))
    body, _ = exports.payout(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=str,
    )

    xml = _sheet_xml(body)
    assert f"<v>{PRECISE_AMOUNT}</v>" in xml
    assert "<v>1234.57</v>" not in xml


def test_payout_money_cell_keeps_format_and_is_read_back_as_a_number():
    """Формат `#,##0.00` не должен потеряться, а ячейка не должна стать текстом."""
    view = slice_cells(list(MATERIAL))
    body, _ = exports.payout(
        view, tenant_id=None, period=date(2026, 6, 1), title="Июнь 2026",
        ledger_title=str,
    )

    book = openpyxl.load_workbook(io.BytesIO(body))
    ws = book.active

    money_cells = [
        c for row in ws.iter_rows(min_row=4) for c in row
        if c.value is not None and c.number_format == exports.MONEY
    ]
    assert money_cells, "в ведомости не нашлось ни одной денежной ячейки"
    for c in money_cells:
        assert not isinstance(c.value, str), f"ячейка {c.coordinate} стала текстом"
        assert Decimal(str(c.value)) == AMOUNT
