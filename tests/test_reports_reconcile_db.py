"""Сверка и выгрузки на живых данных, через страницы (T031, T032).

Что здесь проверяется и почему именно так.

**Границу видимости держит база, а не приложение.** Поэтому все проверки идут
через страницу под конкретной ролью, а контрольные суммы снимаются с базы
**ролью `app_user`**: владелец таблиц и суперпользователь обходят RLS, и под
ними сошлось бы даже при снятых политиках.

**Выгрузка читается обратно.** Файл открывается openpyxl, числа складываются и
сравниваются с тем, что показано на экране. «Ответ вернулся» не означает
ничего: файл может быть пустым или собранным из другой выборки.

**Файл роли с ограниченным доступом не знает о чужих регистрах** — ни строкой,
ни названием, ни вкладом в итог (D023). Это главная проверка T032: выгрузка
уходит из продукта и живёт своей жизнью.
"""
from __future__ import annotations

import io
import re
from decimal import Decimal

import openpyxl
import pytest

from conftest import PLATA_SAMPLE, body, login_as, narrowed_ledgers, period_url, wipe_payruns

JUNE = "2026-06-01"

# Ориентиры приёмки, снятые на данных сида (те же, что у ведомости, T028).
#
# У бухгалтера после D036 набор регистров полон, как у директора (равный
# доступ), и её точки в сиде тоже не сужены — поэтому её число теперь равно
# директорскому, а не отдельно снятому срезу. Ниже, где проверка именно про
# роль с неполным набором, а не про бухгалтера как таковую, набор сужается
# явно `narrowed_ledgers` — это единственная роль в сиде с правом
# `payrun.calculate` и правом видеть все точки, поэтому воспроизвести старый
# срез («видит всё, кроме части регистров») можно только на ней.
CONTROL = {
    "director": Decimal("1951806.13"),
    "accountant": Decimal("1951806.13"),
    "manager": Decimal("891373.32"),
}

# Набор, до которого сужается бухгалтер там, где тест именно про роль с
# неполным набором (см. комментарий к CONTROL выше). Значение то же, что был
# её умолчательный набор до D036 — так контрольные суммы ниже не меняются.
NARROWED_ACCOUNTANT = ["official"]

HIDDEN_FROM_ACCOUNTANT = ("Дополнительный", "Внутренний", "supplementary", "internal")

# У управляющего свой набор: два регистра из трёх. Список отдельный, а не
# срезанный с бухгалтерского, потому что «чего не видно» — свойство роли, и
# общая константа однажды объявила бы утечкой его собственные данные (на чём
# этот тест и попался при написании).
#
# Названий чужих точек здесь нет намеренно: на странице сверки они приходят с
# листов файла, который человек загрузил **сам**. Продукт о них не сообщает
# ничего — он читает принесённое. Требовать их отсутствия значило бы запретить
# управляющему открыть таблицу бухгалтера, то есть сверку целиком.
HIDDEN_FROM_MANAGER = ("Внутренний", "internal")


@pytest.fixture
def calculated_june(client, web_env):
    """Посчитанный июнь на данных сида — общий материал для проверок ниже."""
    wipe_payruns(web_env)
    login_as(client, "director")
    response = client.post(period_url(client) + "calculate/", follow=True)
    assert response.status_code == 200
    return None


def reconcile_page(client, user: str) -> str:
    """Загрузить обезличенную таблицу на страницу сверки — как это делает человек."""
    login_as(client, user)
    with PLATA_SAMPLE.open("rb") as handle:
        response = client.post(
            period_url(client) + "reconcile/", {"table": handle}, follow=True
        )
    assert response.status_code == 200, f"{user}: сверка ответила {response.status_code}"
    return body(response)


def download(client, user: str, kind: str, query: str = ""):
    login_as(client, user)
    response = client.get(period_url(client) + f"export/{kind}/{query}")
    assert response.status_code == 200, f"{user}/{kind}: ответ {response.status_code}"
    assert "spreadsheetml" in response["Content-Type"], "это не книга Excel"
    assert "attachment" in response["Content-Disposition"]
    return openpyxl.load_workbook(io.BytesIO(response.content))


def text_of(book) -> str:
    return "\n".join(
        str(value)
        for ws in book.worksheets
        for row in ws.iter_rows(values_only=True)
        for value in row
        if value is not None
    )


def column(ws, header: str) -> list:
    """Значения колонки по её заголовку — как её находит глазами человек."""
    index = None
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(v) if v is not None else None for v in row]
        if index is not None and row[index] is not None:
            out.append(row[index])
        if header in cells:
            index = cells.index(header)
    assert index is not None, f"в листе «{ws.title}» нет колонки «{header}»"
    return out


SUMMARY_ROW = re.compile(r"<td>([^<]+)</td>\s*<td class=\"num[^\"]*\">([^<]+)</td>")


def summary(html: str) -> dict[str, int | str]:
    """Сводка сверки значениями, а не наличием подписи.

    Подписи в сводке стоят всегда, включая нулевые строки: это её смысл —
    человек должен видеть, что потерянных строк ноль, а не догадываться об
    этом по отсутствию раздела. Значит проверка «подпись есть на странице»
    не проверяет ничего и зеленеет при любом содержимом. Здесь читаются значения.

    Значение не всегда число: строка про расчёт целиком у роли, которой расчёт
    отдан не весь, говорит словом «не проверялось» (T100). Читается оно тем же
    разбором намеренно — счёт строк сводки остаётся сторожем: раньше выражение
    требовало цифр, и такая строка молча выпала бы из сводки вместе с проверкой
    её полноты.
    """
    start = html.index("<th>Итог сверки</th>")
    # Только тело таблицы: в подвале стоит сумма к выплате, и она тоже
    # `<td class="num">`. Пока выражение требовало цифр, подвал отсеивался сам
    # собой — с приходом словесного значения (T100) он попал бы в сводку и
    # сломал счёт её строк.
    table = html[html.index("<tbody>", start):html.index("</tbody>", start)]
    found = {
        label.strip(): int(value) if value.strip().isdigit() else value.strip()
        for label, value in SUMMARY_ROW.findall(table)
    }
    assert len(found) == 6, f"сводка сверки прочитана не целиком: {found}"
    return found


CELL = re.compile(r"<td[^>]*>(.*?)</td>", re.S)
TAG = re.compile(r"<[^>]+>")


def section_rows(html: str, heading: str) -> list[list[str]]:
    """Строки таблицы под заголовком — то, что человек действительно видит.

    Сводку считает ядро сверки, а разделы раскладывает шаблон, и разойтись они
    могут молча: сводка скажет «разошлось 0», а на экране будут висеть те же
    строки в разделе «Разошлось» — без единого числа, а значит и без имени,
    потому что имя стоит в первой ячейке первого числа. Проверено порчей:
    проверка одной сводки такую перестановку не ловит.
    """
    start = html.index(f"<h2>{heading}")
    table = html[start:html.index("</table>", start)]
    body_at = table.index("<tbody>")
    return [
        [TAG.sub(" ", cell).strip() for cell in CELL.findall(row)]
        for row in table[body_at:].split("<tr")[1:]
    ]


def withheld_totals(dsn: str) -> set[str]:
    """Итоги расчёта так, как их напечатала бы страница, — глазами директора.

    Ровно эти строки не имеет права встретиться на сверке роли, которой база
    итогов не отдала: совпадение означало бы, что скрытая часть уехала на
    экран. Снимаются ролью `app_user` под директором, а не владельцем схемы:
    владельца политики не касаются, и набор был бы тот же при снятых.
    """
    import psycopg

    from conftest import as_app_user
    from core.db_types import register_enum_types
    from core.management.commands.seed_dev import det_id
    from web.format import money

    with psycopg.connect(dsn) as conn:
        register_enum_types(conn)
        with as_app_user(conn, str(det_id("user", "director"))) as scoped:
            rows = scoped.execute(
                """select t.net, t.gross, t.contributions, t.total_cost
                     from payslip_totals t
                     join payslips p on p.id = t.payslip_id
                     join payruns r on r.id = p.payrun_id
                    where r.period = %s""",
                (JUNE,),
            ).fetchall()

    # Ноль отбрасывается намеренно: он встречается в разметке сам по себе и
    # объявил бы утечкой любую страницу.
    return {money(value) for row in rows for value in row if value}


def visible_total(dsn: str, user: str) -> Decimal:
    """Сколько база отдаёт этой роли — её же ролью, а не владельцем схемы."""
    import psycopg

    from conftest import as_app_user
    from core.db_types import register_enum_types
    from core.management.commands.seed_dev import det_id

    with psycopg.connect(dsn) as conn:
        register_enum_types(conn)
        with as_app_user(conn, str(det_id("user", user))) as scoped:
            return scoped.execute(
                """select coalesce(sum(c.amount), 0)
                     from pay_components c
                     join payslips p on p.id = c.payslip_id
                     join payruns r on r.id = p.payrun_id
                    where r.period = %s""",
                (JUNE,),
            ).fetchone()[0]


# --- сверка ------------------------------------------------------------------


def test_the_reconciliation_of_the_reference_table_matches(client, calculated_june):
    """Эталонная таблица июня сходится с расчётом продукта построчно.

    Настоящая таблица бухгалтерии в репозиторий не попадает никогда (D028), и
    её загрузка — приёмочный шаг MVP (T039). Здесь тот же формат и те же четыре
    схемы на обезличенном файле: сверка обязана сойтись и назвать всё, что не
    сошлось.
    """
    html = reconcile_page(client, "director")
    counts = summary(html)

    assert "Сверка с таблицей бухгалтера" in html
    assert "Не разобрано в файле" not in html, "обезличенный файл разбирается целиком"
    assert counts["Сошлось до копейки"] == 32, (
        f"не все 32 строки эталонной таблицы сошлись с расчётом: {counts}"
    )
    assert counts["Разошлось"] == 0
    assert counts["Разошлось на копейки (округление)"] == 0
    assert counts["Есть в таблице, нет в вашей части расчёта"] == 0, (
        "директору видны все строки — терять их сверке негде"
    )
    assert counts["Сверены только входы — деньги не сравнивались"] == 0, (
        "директору отданы все итоги: сверять по одним входам ему нечего"
    )
    # Курьеры и строка исправления есть в сиде и не могут быть в таблице
    # бухгалтера: она за другой месяц другой сети. Это не потеря, а разница
    # наборов, и она названа числом, чтобы молчаливый рост был виден.
    assert counts["Есть в расчёте, нет в таблице"] == 3, (
        f"в расчёте сида ровно три человека сверх таблицы: {counts}"
    )


def test_the_reconciliation_never_reports_a_match_it_did_not_make(
    client, web_env, calculated_june
):
    """Главная проверка T031 под ролью со скрытыми итогами.

    Нето, бруто и взносы посчитаны по строке ведомости целиком, поэтому база
    отдаёт их только роли, которой видны все регистры (T071). У бухгалтера
    после D036 они есть — её набор полон, как у директора, — поэтому сюда
    нужна роль с её же правами, но урезанным набором: он собирается явно
    `narrowed_ledgers` (см. комментарий у `CONTROL`), а не берётся из умолчаний
    сида, которых для такого сочетания прав больше нет.

    Ловушка, ради которой тест написан: пустое `all()` даёт истину. Строка, в
    которой нечего было сравнивать, без различения «сравнивали» и «сошлось»
    попадала бы в сошедшиеся, и роль прочитала бы «сошлось до копейки» по
    расчёту, которого не видела.
    """
    with narrowed_ledgers(web_env, "accountant", NARROWED_ACCOUNTANT):
        html = reconcile_page(client, "accountant")
    counts = summary(html)

    assert counts["Сверены только входы — деньги не сравнивались"] == 32, (
        f"итоги бухгалтеру не отданы — сверять по деньгам нечего: {counts}"
    )
    assert counts["Сошлось до копейки"] == 0, (
        f"сверка объявила совпадением строки, деньги которых не сравнивала: {counts}"
    )
    assert counts["Разошлось"] == 0, (
        f"несравнённое — не расхождение: сравнивать было нечего: {counts}"
    )
    assert "Всё сошлось до копейки" not in html, (
        "сверка назвала себя чистой, не сверив ни одной суммы"
    )
    assert "Деньги не сравнивались ни по одной строке" in html, (
        "подвал обязан сказать, что суммы не сверялись, а не показать ноль"
    )

    # Сводка и разделы считаются в разных местах и расходятся молча. Раздел
    # «Разошлось» с этими строками — не косметика: имя в нём стоит в ячейке
    # первого показанного числа, а чисел по такой строке нет ни одного, и
    # бухгалтер получил бы 32 безымянные строки с подписью про правило расчёта.
    assert "<h2>Разошлось</h2>" not in html, (
        "несравнённые строки показаны как расхождение"
    )
    rows = section_rows(html, "Сверены только входы")
    assert len(rows) == 32, f"в разделе {len(rows)} строк вместо 32"
    for sheet, name, what in rows:
        assert name, f"строка раздела без имени: {sheet!r} {what!r}"
        assert what, f"строка {name!r} ничего не говорит про входы"


def test_the_reconciliation_never_shows_our_numbers_for_a_row_it_cannot_see(
    client, web_env, calculated_june
):
    """Главная проверка D023 на сверке: скрытый регистр не выдаётся вычитанием.

    Итог файла минус сумма видимых компонентов — это ровно скрытая часть.
    Поэтому по строке без отданных итогов не показывается **ни одного нашего
    числа**: ни суммы, ни разницы. Проверяется самими числами расчёта, а не
    наличием раздела на странице.

    Роль с реально урезанным набором собирается явно (см. `CONTROL`): у
    бухгалтера после D036 набор полон, а у управляющего сужен вместе с точкой,
    и совпавшую с этим тестом форму — «все точки, часть регистров» — в сиде
    больше никто не даёт.
    """
    with narrowed_ledgers(web_env, "accountant", NARROWED_ACCOUNTANT):
        html = reconcile_page(client, "accountant")

        for word in HIDDEN_FROM_ACCOUNTANT:
            assert word not in html, f"на странице сверки бухгалтера есть «{word}»"

    hidden = withheld_totals(web_env)
    assert len(hidden) >= 30, f"нечего проверять: итогов в расчёте {len(hidden)}"
    for shown in hidden:
        assert shown not in html, (
            f"на сверке бухгалтера стоит наш итог {shown} по строке, "
            f"итоги которой база ему не отдала"
        )


def test_the_reconciliation_does_not_pass_off_invisible_as_absent(
    client, calculated_june
):
    """T095 на живых данных: управляющему не сообщают о лишних людях в таблице.

    У управляющего одна точка, и строки остальных точек база ему не отдаёт.
    Сверка про них говорила «в расчёте этого периода такой строки нет» — то
    есть выдавала свою границу видимости за факт о базе. Для управляющего это
    читалось как 16 лишних человек в таблице бухгалтера.

    Проверяется тремя вещами сразу: строки по-прежнему найдены и посчитаны (не
    выброшены), формулировка ничего не отрицает — и ни в ней, ни в заголовке
    раздела не названо ни одной чужой точки и ни одного регистра.
    """
    html = reconcile_page(client, "manager")
    counts = summary(html)

    missing = counts["Есть в таблице, нет в вашей части расчёта"]
    assert missing == 16, (
        f"на данных сида управляющему не видны ровно 16 строк таблицы: {counts}"
    )

    rows = section_rows(html, "Есть в таблице, нет в вашей части расчёта")
    assert len(rows) == missing, f"в разделе {len(rows)} строк вместо {missing}"
    for _sheet, name, why in rows:
        assert name, "строка раздела без имени"
        assert "нет" not in why, (
            f"сверка отрицает строку, которой не проверяла: {why!r} ({name})"
        )

    # Управляющему видны официальный и дополнительный регистры и одна точка.
    # Скрыто от него другое — и починка формулировки не имеет права это назвать.
    for word in HIDDEN_FROM_MANAGER:
        assert word not in html, (
            f"сверка управляющего называет «{word}» — то, чего ему не видно"
        )


def test_the_reconciliation_still_says_plainly_what_the_file_lacks(
    client, calculated_june
):
    """Обратная сторона не смягчается: файл человек видит целиком.

    Ловушка при починке T095 — заодно размыть и это. «В загруженной таблице
    такой строки нет» проверено загруженным файлом, и мягчить проверенный факт
    значит терять его.
    """
    html = reconcile_page(client, "director")
    rows = section_rows(html, "Есть в расчёте, нет в таблице")

    assert len(rows) == 3, f"в разделе {len(rows)} строк вместо трёх"
    for _name, why in rows:
        assert "нет" in why, f"факт о загруженном файле размыт: {why!r}"


def test_the_reconciliation_refuses_a_file_that_is_not_a_workbook(client, calculated_june):
    login_as(client, "director")
    response = client.post(
        period_url(client) + "reconcile/",
        {"table": io.BytesIO(b"nonsense")},
        follow=True,
    )
    assert response.status_code == 422
    assert "не удалось прочитать" in body(response)


def test_the_reconciliation_says_when_no_file_was_chosen(client, calculated_june):
    login_as(client, "director")
    response = client.post(period_url(client) + "reconcile/", {}, follow=True)
    assert response.status_code == 400
    assert "Файл не выбран" in body(response)


# --- выгрузки ----------------------------------------------------------------


@pytest.mark.parametrize("kind", ["payout", "pnl", "partner"])
def test_every_export_opens_as_a_workbook_with_data(client, calculated_june, kind):
    book = download(client, "director", kind)

    assert book.worksheets, f"{kind}: в книге нет ни одного листа"
    assert "Июнь 2026" in text_of(book), f"{kind}: в файле не сказано, какой это месяц"


def test_the_payout_export_carries_the_same_total_the_role_sees(
    client, web_env, calculated_june
):
    """Файл сходится с базой, спрошенной ролью `app_user`, — для каждой роли."""
    for user, expected in CONTROL.items():
        book = download(client, user, "payout")
        *lines, footer = [Decimal(str(v)) for v in column(book.active, "Итого")]

        assert sum(lines, Decimal(0)) == footer, f"{user}: подвал файла не равен его строкам"
        assert footer == expected, f"{user}: в файле {footer}, ожидалось {expected}"
        assert footer == visible_total(web_env, user), (
            f"{user}: файл показывает не то, что отдаёт база его же ролью"
        )


def test_the_payout_export_repeats_the_screen_row_for_row(client, calculated_june):
    """Скачанный файл сверяется **с экраном**, а не сам с собой.

    «Итог файла равен сумме его строк» держит файл в согласии с самим собой и
    молчит, когда выгрузка собрала другую выборку. Здесь книга читается обратно
    и сравнивается со страницей, с которой человек её скачал: те же строки, в
    том же порядке, на те же суммы — для каждой роли и в каждом её разрезе.

    Читает страницу тот же разборщик, что и проверки ведомости (T028): вторые
    «глаза» рядом с первыми разъехались бы с ними молча.
    """
    from test_reports_sheet import page, row_totals, shown_total

    for user in CONTROL:
        for query in ("", "?ledger=official"):
            screen = page(client, user, query)
            book = download(client, user, "payout", query)
            *lines, footer = [Decimal(str(v)) for v in column(book.active, "Итого")]

            assert lines == row_totals(screen), (
                f"{user}{query}: строки файла не совпадают со строками экрана"
            )
            assert footer == shown_total(screen), (
                f"{user}{query}: подвал файла {footer}, а на экране "
                f"{shown_total(screen)}"
            )


@pytest.mark.parametrize("kind", ["payout", "pnl", "partner"])
def test_no_export_of_the_accountant_knows_about_other_ledgers(
    client, web_env, calculated_june, kind
):
    """Главная проверка T032: чужого регистра в файле нет ни в каком виде.

    После D036 у бухгалтера набор регистров полон, и по умолчанию она про
    другие регистры знает законно. Роль с её же правами и урезанным набором
    собирается явно `narrowed_ledgers` (см. `CONTROL`) — так же, как выше на
    сверке того же файла.
    """
    with narrowed_ledgers(web_env, "accountant", NARROWED_ACCOUNTANT):
        text = text_of(download(client, "accountant", kind))

        for word in HIDDEN_FROM_ACCOUNTANT:
            assert word not in text, f"{kind}: в файле бухгалтера есть «{word}»"


def test_the_export_of_a_cut_holds_exactly_that_cut(client, calculated_june):
    """Выгрузка берёт тот же разрез, что показан на экране, а не всю ведомость."""
    parts = {}
    for ledger in ("official", "supplementary", "internal"):
        book = download(client, "director", "payout", f"?ledger={ledger}")
        *lines, footer = [Decimal(str(v)) for v in column(book.active, "Итого")]
        assert sum(lines, Decimal(0)) == footer
        parts[ledger] = footer

    assert sum(parts.values(), Decimal(0)) == CONTROL["director"], (
        "сумма разрезов не сходится с полной ведомостью"
    )


def test_a_cut_the_role_cannot_see_gives_the_same_file_as_no_cut(
    client, web_env, calculated_june
):
    """Подобранный `?ledger=internal` у роли, которой он не виден, — её обычная выгрузка.

    После D036 бухгалтеру внутренний регистр виден, и адрес перестал быть для
    неё подобранным. Роль с урезанным набором собирается явно `narrowed_ledgers`
    (см. `CONTROL`), чтобы проверка снова стояла на настоящем ограничении.
    """
    with narrowed_ledgers(web_env, "accountant", NARROWED_ACCOUNTANT):
        guessed = download(client, "accountant", "payout", "?ledger=internal")
        plain = download(client, "accountant", "payout")

        assert text_of(guessed) == text_of(plain)


def test_the_pnl_export_splits_accruals_from_taxes(client, calculated_june):
    """Начисления сходятся с ведомостью, налоги стоят отдельными строками."""
    book = download(client, "director", "pnl")
    ws = book.active

    kinds = column(ws, "Тип строки")
    amounts = column(ws, "Сумма")
    accruals = [
        Decimal(str(amount))
        for kind, amount in zip(kinds, amounts, strict=True)
        if kind == "Начисление"
    ]

    assert sum(accruals, Decimal(0)) == CONTROL["director"]
    assert "Налог" in kinds and "Взносы" in kinds


def test_an_unknown_export_is_not_found(client, calculated_june):
    login_as(client, "director")
    assert client.get(period_url(client) + "export/secret/").status_code == 404
