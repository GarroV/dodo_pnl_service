"""
Импорт таблицы партнёра в табель (T020) и отчёт разбора (T021).

Три уровня, и каждый ловит своё:

1. **Разбор файла** — без базы. Здесь проверяется не то, что эталон читается
   (это делает регрессия движка), а то, что **испорченный файл не молчит**:
   чужой лист, переименованная колонка, текст вместо числа и непронумерованная
   строка обязаны попасть в находки, а не пропасть.
2. **Загрузка в продукт** — на живой базе. Эталонная таблица даёт корректный
   табель, суммы расчёта после неё те же, а повторная загрузка не меняет
   **ничего**: ни строк, ни чисел, ни идентификаторов дней.
3. **Проверки данных** — часов больше нормы, сотрудник без часов, часы у
   уволенного: подсказки, ради которых отчёт и нужен.
"""
from __future__ import annotations

import re
import shutil
from datetime import date
from decimal import Decimal

import pytest

from conftest import PLATA_SAMPLE, wipe_payruns

JUNE = date(2026, 6, 1)


# =============================================================================
# 1. Разбор файла: находки вместо молчания
# =============================================================================


@pytest.fixture
def sample_copy(tmp_path):
    """Копия эталона, которую можно портить."""
    def make(name: str = "broken.xlsx"):
        target = tmp_path / name
        shutil.copy(PLATA_SAMPLE, target)
        return target
    return make


def _read(path):
    from payroll.importers import read_plata_file

    return read_plata_file(path)


def test_clean_reference_has_no_findings():
    """Эталон обязан читаться начисто: иначе находки перестанут что-то значить."""
    parsed = _read(PLATA_SAMPLE)
    assert len(parsed.rows) == 32
    assert parsed.findings == [], [f.text for f in parsed.findings]


def test_reader_keeps_reading_the_same_rows_as_before():
    """Старый вход не должен поехать: на нём стоит вся регрессия движка."""
    from payroll.importers import read_plata, read_plata_file

    assert [row.name for row in read_plata(PLATA_SAMPLE)] == [
        row.name for row in read_plata_file(PLATA_SAMPLE).rows
    ]


def test_unknown_sheet_is_reported(sample_copy):
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    wb.create_sheet("Заметки бухгалтера")
    wb.save(path)

    parsed = _read(path)
    assert len(parsed.rows) == 32, "чужой лист не должен уносить разобранные строки"
    assert [f.where for f in parsed.findings if f.kind == "sheet"] == [
        "Заметки бухгалтера"
    ]


def test_missing_sheet_of_the_format_is_reported(sample_copy):
    """Листа формата нет — значит его людей в загрузке нет. Это надо сказать."""
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    del wb["NS 2 Dunavska "]
    wb.save(path)

    parsed = _read(path)
    assert len(parsed.rows) == 28
    missing = [f for f in parsed.findings if f.kind == "sheet"]
    assert len(missing) == 1
    assert "NS 2 Dunavska" in missing[0].where


def test_renamed_column_is_reported(sample_copy):
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip().upper().startswith("SATI RADA"):
            ws.cell(1, col).value = "SATI (novo)"
            break
    else:  # pragma: no cover — эталон обязан содержать эту колонку
        pytest.fail("в эталоне нет колонки часов — тест проверял бы не то")
    wb.save(path)

    parsed = _read(path)
    columns = [f for f in parsed.findings if f.kind == "column"]
    assert columns, "переименованная колонка прошла молча"
    assert any("regular" in f.text or "SATI RADA" in f.text for f in columns)


def test_text_instead_of_number_is_reported_and_row_is_not_guessed(sample_copy):
    """Худший исход — принять «восемь» за ноль и посчитать зарплату по нему."""
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip().upper().startswith("SATI RADA"):
            ws.cell(2, col).value = "восемь"
            break
    wb.save(path)

    parsed = _read(path)
    assert len(parsed.rows) == 31, "строка с нечисловым значением не должна попасть в загрузку"
    values = [f for f in parsed.findings if f.kind == "value"]
    assert values and "восемь" in values[0].text


def test_unnumbered_row_with_a_name_is_reported(sample_copy):
    """Раньше такие строки исчезали молча — вместе с человеком."""
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    ws.cell(6, 2).value = "ZABORAVLJENI"
    ws.cell(6, 3).value = "RADNIK"
    wb.save(path)

    parsed = _read(path)
    assert len(parsed.rows) == 32
    rows = [f for f in parsed.findings if f.kind == "row"]
    assert rows and "ZABORAVLJENI" in rows[0].text


def test_rows_below_the_old_scan_limit_are_read(sample_copy):
    """Разбор ограничивался 18 строками на лист: 19-й человек пропадал молча."""
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    for line in range(6, 26):
        for col in range(1, ws.max_column + 1):
            ws.cell(line, col).value = ws.cell(2, col).value
        ws.cell(line, 1).value = line
        ws.cell(line, 2).value = f"DVADESET{line}"
    wb.save(path)

    parsed = _read(path)
    assert len(parsed.rows) == 52


# =============================================================================
# 2. Загрузка в продукт
# =============================================================================


# Снимок табеля и возврат к нему (`period_restored`) живёт в `conftest.py`: он
# нужен не только импорту, а каждому, кто пишет в табель общей базы. Второй
# экземпляр здесь означал бы две правды об одном правиле — и одна из них молча
# отстала бы.


def _tenant_id():
    from core.models import Tenant

    return Tenant.objects.get(code="rs-dev").id


def _import(path=PLATA_SAMPLE, **kw):
    from timesheets.importer import import_partner_table

    with open(path, "rb") as handle:
        return import_partner_table(handle, tenant_id=_tenant_id(), period=JUNE, **kw)


def test_import_of_reference_table_fills_the_timesheet(period_restored):
    """Строки табеля заводятся заново и повторяют числа файла."""
    from core.models import Timesheet
    from payroll.importers import read_plata

    Timesheet.objects.filter(period=JUNE).delete()

    result = _import()
    assert result.created == 32
    assert result.matched == 32
    assert result.unmatched_rows == []

    file_rows = {row.employee.ext_id: row for row in read_plata(PLATA_SAMPLE)}
    stored = Timesheet.objects.select_related("employee").filter(period=JUNE)
    assert stored.count() == 32
    for row in stored:
        source = file_rows[row.employee.external_id]
        assert row.insured_hours == source.timesheet.insured_hours
        for kind, hours in source.timesheet.hours.items():
            assert Decimal(str(row.hours.get(kind, 0))) == hours, (row.employee.external_id, kind)


def test_second_import_changes_nothing(period_restored):
    """Ни строк, ни чисел, ни идентификаторов дней (D023, критерий T020)."""
    from core.models import Timesheet, TimesheetDay

    Timesheet.objects.filter(period=JUNE).delete()
    _import()
    assert TimesheetDay.objects.filter(timesheet__period=JUNE).exists(), (
        "проверять неизменность дней имеет смысл только когда они есть"
    )

    def snapshot():
        sheets = sorted(
            (str(r.id), str(r.employee_id), str(r.unit_id), str(r.hours),
             str(r.insured_hours), str(r.norm_hours), str(r.deduction),
             str(r.cash_payout), str(r.manual_correction), r.created_at.isoformat())
            for r in Timesheet.objects.filter(period=JUNE)
        )
        days = sorted(
            (str(d.id), str(d.timesheet_id), d.work_date.isoformat(), d.hour_type,
             str(d.hours), d.source, d.created_at.isoformat())
            for d in TimesheetDay.objects.filter(timesheet__period=JUNE)
        )
        return sheets, days

    before = snapshot()
    again = _import()

    assert again.created == 0
    assert again.updated == 0
    assert again.unchanged == 32
    assert snapshot() == before


def test_import_over_edited_row_keeps_days_in_sync(period_restored):
    """Строка уже подневная — импорт обязан переписать и дни, а не только итог.

    Иначе инвариант «итог равен сумме дней» рвётся ровно там, где табель уже
    правили руками, — и заметить это стало бы нечем.
    """
    from core.models import Timesheet
    from timesheets import store

    row = (
        Timesheet.objects.select_related("employee")
        .filter(period=JUNE, employee__external_id="VUK MILOSEVIC")
        .first()
    )
    store.set_cell(timesheet=row, hour_type="regular", hours=Decimal("11.00"))
    assert store.daily_totals(row)["regular"] == Decimal("11.00")

    result = _import()
    assert result.updated >= 1

    row.refresh_from_db()
    totals = store.daily_totals(row)
    assert totals["regular"] == Decimal(str(row.hours["regular"]))
    assert totals["regular"] != Decimal("11.00")


def test_import_does_not_move_the_calculation(web_env, period_restored):
    """Загрузка эталона поверх сида не двигает суммы месяца ни на копейку.

    Импорт пишет тот же файл, из которого собран сид, поэтому сдвинуться нечему —
    и если сдвинулось, то не в данных, а в записи.

    **Сравнивается с итогом, снятым здесь же до загрузки, а не с зашитым
    числом** (issue #69). Пока в обеих проверках стояло 1 951 806,13, тест
    сторожил не импорт, а состояние общей базы: любой сосед, оставивший
    изменённой ставку, норму часов или версию правила, красил его — при
    исправном импорте. Красный не по своей вине хуже отсутствующего теста:
    настоящую поломку в следующий раз спишут на тот же шум. Хуже того, зелёным
    он был по стечению обстоятельств — из-за порядка файлов в прогоне, — то
    есть не проверял ничего.

    Ориентир приёмки (60 строк ведомости директора на 1 951 806,13) при этом не
    потерян: его сторожат те, кто владеет этим состоянием, —
    `test_norm_hours.JUNE_TOTAL`, `test_payslip_freezing`, `test_reports_sheet`.
    """
    from django.db.models import Sum

    from core.models import PayComponent
    from payrun.calc import calculate_period

    all_ledgers = ["official", "supplementary", "internal"]

    def total_of_june() -> Decimal:
        wipe_payruns(web_env)
        run = calculate_period(tenant_id=_tenant_id(), period=JUNE, visible_ledgers=all_ledgers)
        return PayComponent.objects.filter(payslip__payrun_id=run.payrun_id).aggregate(
            total=Sum("amount")
        )["total"]

    before = total_of_june()
    assert before, "июнь не посчитался вовсе — сравнивать после загрузки будет не с чем"

    _import()

    assert total_of_june() == before


def test_unknown_employee_lands_in_unmatched_rows(period_restored):
    """Человека нет в справочнике — строка не теряется, а называется.

    Сотрудник не удаляется, а переименовывается: удаление унесло бы каскадом и
    его табель, и условия найма, и тест чинил бы базу вместо проверки.
    """
    from core.models import Employee

    Employee.objects.filter(external_id="VUK MILOSEVIC").update(external_id="ушёл-в-никуда")
    try:
        result = _import()
        assert [row.name for row in result.unmatched_rows] == ["VUK MILOSEVIC"]
        assert result.matched == 31
    finally:
        Employee.objects.filter(external_id="ушёл-в-никуда").update(
            external_id="VUK MILOSEVIC"
        )


def test_rows_of_a_closed_unit_are_reported_not_written(period_restored):
    """Закрытая точка (T022) не загружается — и говорит об этом словами.

    Без этой ветки загрузка упиралась бы в политику базы и падала целиком:
    транзакция у импорта одна на файл, поэтому одна закрытая точка уносила бы и
    все остальные строки. Человек при этом прочитал бы «файл не удалось
    прочитать как книгу Excel» — неправду о причине.
    """
    from core.models import Timesheet, TimesheetClosure, Unit

    unit = Unit.objects.get(tenant__code="rs-dev", code="NS1")
    before = {
        row.employee.external_id: dict(row.hours or {})
        for row in Timesheet.objects.select_related("employee").filter(
            period=JUNE, unit=unit
        )
    }
    assert before, "в сиде нет строк точки NS1 — тест бессмысленен"

    closure = TimesheetClosure.objects.create(
        tenant_id=_tenant_id(), unit=unit, period=JUNE
    )
    try:
        result = _import()
        refused = [row.why for row in result.skipped_rows]
        assert refused, "закрытая точка загрузилась молча"
        assert all("закрыт" in why for why in refused), refused
        # T122: продукт назвал этих людей поимённо — значит сопоставил, и
        # звать их несопоставленными он не вправе.
        assert result.unmatched_rows == [], (
            "строки закрытой точки снова попали в «не сопоставлено», хотя "
            f"сопоставлены: {[row.name for row in result.unmatched_rows]}"
        )
    finally:
        closure.delete()

    after = {
        row.employee.external_id: dict(row.hours or {})
        for row in Timesheet.objects.select_related("employee").filter(
            period=JUNE, unit=unit
        )
    }
    assert after == before


# =============================================================================
# 3. Проверки данных: подсказки о подозрительном (T021)
# =============================================================================


def _flagged(result, kind: str) -> list[str]:
    """Кого именно подсветила загрузка — по ключу сотрудника, не по имени."""
    return [note.external_id for note in result.warnings if note.kind == kind]


def test_hours_above_norm_are_flagged(period_restored, sample_copy):
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip().upper().startswith("SATI RADA"):
            ws.cell(2, col).value = 400
            break
    wb.save(path)

    result = _import(path)
    assert "VUK MILOSEVIC" in _flagged(result, "over_norm")


def test_employee_without_hours_is_flagged(period_restored, sample_copy):
    import openpyxl

    path = sample_copy()
    wb = openpyxl.load_workbook(path)
    ws = wb["NS 1 Bulevar "]
    for col in range(1, ws.max_column + 1):
        title = str(ws.cell(1, col).value or "").strip().upper()
        if title.startswith(("SATI RADA", "SAT NA RAD", "GODISNJI", "SATI BOLOVANJE")):
            ws.cell(2, col).value = 0
    wb.save(path)

    result = _import(path)
    assert "VUK MILOSEVIC" in _flagged(result, "no_hours")


def test_hours_of_a_dismissed_employee_are_flagged(period_restored):
    from core.models import Employee

    Employee.objects.filter(external_id="VUK MILOSEVIC").update(
        dismissed_at=date(2026, 5, 31)
    )
    try:
        result = _import()
        assert "VUK MILOSEVIC" in _flagged(result, "dismissed")
    finally:
        Employee.objects.filter(external_id="VUK MILOSEVIC").update(dismissed_at=None)


def test_clean_reference_import_has_no_data_warnings(period_restored):
    """Иначе предупреждения станут фоном, который перестают читать."""
    result = _import()
    assert result.warnings == [], [note.text for note in result.warnings]


# =============================================================================
# 4. Экран загрузки
# =============================================================================


def _grid_url(client) -> str:
    import re

    from conftest import period_url

    match = re.search(r"([0-9a-f-]{36})", period_url(client))
    return f"/timesheets/{match.group(1)}/"


def test_grid_offers_the_upload_to_those_who_may_edit(client, web_env):
    from conftest import body, login_as

    login_as(client, "director")
    html = body(client.get(_grid_url(client)))
    assert 'name="table"' in html, "на табеле нет формы загрузки"


def test_grid_does_not_offer_the_upload_without_the_right(client, web_env):
    """Кнопка, которая заведомо даст 403, — обещание, которого экран не держит."""
    from conftest import body, login_as

    login_as(client, "admin")
    assert 'name="table"' not in body(client.get(_grid_url(client)))


def test_upload_shows_the_report(client, period_restored):
    from conftest import body, login_as

    login_as(client, "director")
    with open(PLATA_SAMPLE, "rb") as handle:
        response = client.post(_grid_url(client) + "import/", {"table": handle})

    assert response.status_code == 200
    html = body(response)
    assert "Сопоставлено сотрудникам" in html
    assert "32" in html


def test_upload_of_a_broken_file_says_so_instead_of_500(client, web_env):
    """Не тот формат — читаемый отказ, а не «Server Error»."""
    import io

    from conftest import body, login_as

    login_as(client, "director")
    fake = io.BytesIO("это не книга Excel".encode())
    fake.name = "notes.xlsx"
    response = client.post(_grid_url(client) + "import/", {"table": fake})

    assert response.status_code == 422
    assert "не удалось прочитать" in body(response)
    assert "Server Error" not in body(response)


def test_upload_without_a_file_is_refused(client, web_env):
    from conftest import body, login_as

    login_as(client, "director")
    response = client.post(_grid_url(client) + "import/", {})
    assert response.status_code == 400
    assert "Файл не выбран" in body(response)


def test_upload_is_refused_without_the_right(client, web_env):
    """Проверка на сервере остаётся, даже когда формы на экране нет."""
    from conftest import body, login_as

    login_as(client, "admin")
    with open(PLATA_SAMPLE, "rb") as handle:
        response = client.post(_grid_url(client) + "import/", {"table": handle})

    assert response.status_code == 403
    assert "Правка табеля" in body(response)


def test_the_report_does_not_call_a_named_row_unmatched(client, period_restored):
    """Счётчик отчёта не вправе врать про класс события (T122).

    Было: «Сопоставлено сотрудникам 16 · Не сопоставлено 16», и тут же список
    этих шестнадцати с причиной «часы точки NS1 за этот месяц закрыты». Их
    сопоставили — их назвали по фамилиям; не загрузили по другой причине.
    Человек читает счётчик раньше списка и уходит искать несуществующее
    расхождение справочника.
    """
    from conftest import body, login_as
    from core.models import TimesheetClosure, Unit

    unit = Unit.objects.get(tenant__code="rs-dev", code="NS1")
    closure = TimesheetClosure.objects.create(
        tenant_id=_tenant_id(), unit=unit, period=JUNE
    )
    login_as(client, "director")
    try:
        with open(PLATA_SAMPLE, "rb") as handle:
            html = body(client.post(_grid_url(client) + "import/", {"table": handle}))
    finally:
        closure.delete()

    unmatched = re.search(
        r"Не сопоставлено</td>\s*<td[^>]*>(\d+)</td>", html, re.S
    )
    assert unmatched, f"в отчёте нет счётчика «Не сопоставлено»:\n{html[:2000]}"
    assert unmatched.group(1) == "0", (
        "строки закрытой точки посчитаны несопоставленными, хотя названы поимённо"
    )
    assert "закрыт" in html, "отчёт перестал говорить, почему строки не загружены"
