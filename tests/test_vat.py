"""НДС у расхода: хранится отдельно, в P&L по умолчанию не показывается (T146, D042).

Ответ владельца на Q016 дословно: «НДС нужен, да. Это вообще важная вещь в
Сербии и вообще. Хотя вообще в итоговом ПНЛ мы обычно не показываем НДС». Из
него следуют ровно четыре правила, и каждое проверено здесь отдельно.

**1. Хранится сумма документа и ставка, а не одна «правильная» сумма.** Человек
вводит то, что написано в чеке, — сумму с налогом, — и ставку. Сумма налога
считается из них и кладётся рядом приколоченной: обратный пересчёт через год
дал бы другую копейку, если поменяется правило округления, а закрытый месяц
двигаться не должен.

**2. В P&L по умолчанию едет сумма БЕЗ НДС.** Умолчание, а не единственный вид:
партнёру, который налог не зачитывает, нужна полная сумма — она достаётся тем же
файлом с явно названным разрезом.

**3. Расход без ставки — законное состояние.** Так внесено всё, что было до этой
задачи, и так вносится всё, где налога нет вовсе. Нетто у такой строки равно
сумме документа, а не нулю.

**4. Разнесение делит и налог.** Иначе сумма без НДС у детей не сошлась бы с
родителем, и разница вылезла бы в P&L копейками, происхождение которых никто не
объяснит.

Проверки идут через экран и через готовый файл выгрузки — то есть тем же путём,
каким числа доходят до бухгалтера. Прямое чтение базы только для осмотра
результата.
"""
from __future__ import annotations

import io
import re
from decimal import Decimal

import openpyxl
import pytest

from conftest import body, login_as
from test_cash_expense import (  # noqa: F401
    JUNE_DAY,
    entry_key,
    facts_removed,
    item,
    payload,
    tenant,
    units,
)
from test_directory import payruns_restored, sql  # noqa: F401
from test_expenses_allocation import rules_removed, set_rule  # noqa: F401

D = Decimal
NEW = "/expenses/new/"
JUNE = "2026-06-01"


def _message(html: str) -> str:
    found = re.search(r'class="alert"[^>]*>(.*?)</div>', html, re.S)
    return re.sub(r"\s+", " ", found.group(1)).strip() if found else ""


def spend(client, item_id, units_map, **extra) -> str:
    """Внести расход июньской датой и вернуть ключ внесения."""
    key = entry_key()
    form = payload(item_id, units_map, entry_key=key, date=JUNE_DAY, **extra)
    answer = client.post(NEW, form)
    assert answer.status_code == 302, body(answer)
    return key


def vat_of(sql, key: str):  # noqa: F811
    """Что записано у факта про деньги и налог — владельцем схемы, это осмотр."""
    return sql.execute(
        """select amount, vat_rate, vat_amount
             from facts where dedup_key = %s and superseded_at is null""",
        (f"manual:cash:{key}",),
    ).fetchall()


# --- 1. сумма документа и ставка живут порознь ---------------------------------


def test_the_expense_keeps_the_document_amount_and_the_rate_apart(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """1200 со ставкой 20% — это 1200 в документе, 200 налога и 1000 без него.

    Сумма факта остаётся суммой документа: её человек видел в чеке, и менять её
    смысл нельзя — на неё смотрят все прежние отчёты.
    """
    login_as(client, "director")
    try:
        key = spend(client, item, units, unit=units["NS1"], amount="1200.00", vat_rate="20")
        assert vat_of(sql, key) == [(D("1200.00"), D("20.000"), D("200.00"))]
    finally:
        client.post("/logout/")


def test_an_expense_without_a_rate_has_no_vat_at_all(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """Пустая ставка — не ноль, а «налога нет»: так внесено всё до этой задачи."""
    login_as(client, "director")
    try:
        key = spend(client, item, units, unit=units["NS1"], amount="500.00")
        assert vat_of(sql, key) == [(D("500.00"), None, None)]
    finally:
        client.post("/logout/")


def test_the_net_amount_of_a_row_without_vat_is_the_whole_amount(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """Без ставки сумма без НДС равна сумме документа, а не нулю.

    Это и есть та ошибка, которая в отчёте выглядит как «половина расходов
    пропала»: `sum(amount - vat_amount)` при пустом `vat_amount` даёт null на
    всю строку, если писать без `coalesce`.
    """
    login_as(client, "director")
    try:
        key = spend(client, item, units, unit=units["NS1"], amount="500.00")
        net = sql.execute(
            """select amount_net from pnl_lines
                where fact_id = (select id from facts
                                  where dedup_key = %s and superseded_at is null)""",
            (f"manual:cash:{key}",),
        ).fetchone()[0]
        assert net == D("500.00"), net
    finally:
        client.post("/logout/")


def test_a_rate_that_cannot_be_a_rate_is_refused_in_words(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """150% и «двадцать» — отказ формы, а не запись мусора и не пятисотка."""
    login_as(client, "director")
    try:
        for wrong in ("150", "-5", "двадцать"):
            answer = client.post(NEW, payload(
                item, units, unit=units["NS1"], amount="100.00",
                vat_rate=wrong, entry_key=entry_key(), date=JUNE_DAY,
            ))
            assert answer.status_code == 400, (wrong, body(answer))
            assert "НДС" in _message(body(answer)), (wrong, _message(body(answer)))
    finally:
        client.post("/logout/")


def test_changing_only_the_rate_counts_as_a_change(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """Смена одной ставки — изменение факта, а не «то же самое событие».

    Без `vat_rate` в `facts_same` повторная отправка с другой ставкой прошла бы
    как `unchanged`, и правка молча не применилась бы.
    """
    login_as(client, "director")
    try:
        key = entry_key()
        form = payload(item, units, unit=units["NS1"], amount="1200.00",
                       entry_key=key, date=JUNE_DAY, vat_rate="20")
        assert client.post(NEW, form).status_code == 302
        assert client.post(NEW, {**form, "vat_rate": "10"}).status_code == 302

        rows = sql.execute(
            """select revision, vat_rate, vat_amount from facts
                where dedup_key = %s and superseded_at is null""",
            (f"manual:cash:{key}",),
        ).fetchall()
        assert rows == [(2, D("10.000"), D("109.09"))], rows
    finally:
        client.post("/logout/")


# --- 2. P&L по умолчанию без НДС ------------------------------------------------


def june_export_url(sql, tenant) -> str:  # noqa: F811
    period_id = sql.execute(
        "select id from periods where tenant_id = %s and period = %s", (tenant, JUNE)
    ).fetchone()[0]
    return f"/periods/{period_id}/export/pnl/"


def expense_rows(client, url) -> list[tuple]:
    response = client.get(url)
    assert response.status_code == 200, response.status_code
    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
    return [
        row for row in sheet.iter_rows(values_only=True)
        if row and row[3] == "Расход"
    ]


@pytest.fixture
def june_expense_with_vat(client, sql, item, units, facts_removed):  # noqa: F811
    """Июньский расход 1200 со ставкой 20% — внесён бухгалтером с экрана."""
    login_as(client, "accountant")
    spend(client, item, units, unit=units["NS1"], amount="1200.00", vat_rate="20")
    client.post("/logout/")


def test_the_pnl_file_shows_the_amount_without_vat_by_default(
    client, sql, tenant, june_expense_with_vat,  # noqa: F811
):
    """Умолчание файла — 1000, а не 1200 (D042).

    Сломайте выбор колонки в `collect_expenses` — и бухгалтер соберёт P&L с
    налогом внутри расходов, не заметив этого: 1200 выглядит ровно так же
    правдоподобно, как 1000.
    """
    login_as(client, "accountant")
    try:
        rows = expense_rows(client, june_export_url(sql, tenant))
        assert [D(str(row[5])) for row in rows] == [D("1000.00")], rows
    finally:
        client.post("/logout/")


def test_the_same_file_can_be_asked_for_the_full_amount(
    client, sql, tenant, june_expense_with_vat,  # noqa: F811
):
    """«Без НДС» — умолчание, а не единственный вид: полная сумма достаётся тем же файлом."""
    login_as(client, "accountant")
    try:
        rows = expense_rows(client, june_export_url(sql, tenant) + "?vat=gross")
        assert [D(str(row[5])) for row in rows] == [D("1200.00")], rows
    finally:
        client.post("/logout/")


def test_the_file_says_which_amounts_it_carries(
    client, sql, tenant, june_expense_with_vat,  # noqa: F811
):
    """В шапке файла написано, с налогом суммы или без.

    Два файла с одинаковыми колонками и разными числами — самый дешёвый способ
    получить P&L, который не сходится, и не понять почему.
    """
    login_as(client, "accountant")
    try:
        url = june_export_url(sql, tenant)
        without = client.get(url).content
        with_vat = client.get(url + "?vat=gross").content
        heads = []
        for raw in (without, with_vat):
            sheet = openpyxl.load_workbook(io.BytesIO(raw)).active
            heads.append(" ".join(
                str(value) for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True)
                for value in row if value
            ))
        assert "без НДС" in heads[0], heads[0]
        assert "с НДС" in heads[1], heads[1]
    finally:
        client.post("/logout/")


def test_the_button_for_the_full_amount_appears_only_when_there_is_vat(
    client, sql, tenant, item, units, facts_removed, payruns_restored,  # noqa: F811
):
    """Кнопки «с НДС» нет, пока НДС не встречается: продукт не обещает пустого.

    Ровно тот же довод, что у налоговой части файла (T141): ряд выгрузок не
    место для кнопки, которая отдаёт то же самое.

    Период здесь считается намеренно: ряд выгрузок появляется под ведомостью, и
    без расчёта на странице нет ни одной кнопки — тест проверял бы не появление
    второй, а отсутствие всех.
    """
    login_as(client, "accountant")
    try:
        period_id = sql.execute(
            "select id from periods where tenant_id = %s and period = %s", (tenant, JUNE)
        ).fetchone()[0]
        assert client.post(
            f"/periods/{period_id}/calculate/", {"inline": "1"}, follow=True
        ).status_code == 200
        page = body(client.get(f"/periods/{period_id}/"))
        assert "Строки для P&amp;L" in page or "Строки для P&L" in page, (
            "ряда выгрузок нет вовсе — проверять появление второй кнопки не на чем"
        )
        assert "vat=gross" not in page, "кнопка «с НДС» есть, а НДС нигде нет"

        spend(client, item, units, unit=units["NS1"], amount="1200.00", vat_rate="20")
        page = body(client.get(f"/periods/{period_id}/"))
        assert "vat=gross" in page, "НДС появился, а достать полную сумму нечем"
    finally:
        client.post("/logout/")


# --- 3. экран расходов показывает налог ------------------------------------------


def test_the_expenses_list_shows_the_vat_of_a_row(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """Список расходов показывает сумму налога, а прочерк — там, где его нет."""
    login_as(client, "director")
    try:
        spend(client, item, units, unit=units["NS1"], amount="1200.00", vat_rate="20")
        spend(client, item, units, unit=units["NS1"], amount="500.00")
        page = body(client.get("/expenses/?from=2026-06-01&to=2026-06-30"))
        assert "200,00" in page or "200.00" in page, "суммы налога не видно"
        assert "НДС" in page, "колонки НДС нет"
    finally:
        client.post("/logout/")


# --- 4. разнесение делит налог ---------------------------------------------------


def test_the_children_of_an_allocation_split_the_vat_too(
    client, sql, tenant, item, units, rules_removed, facts_removed,  # noqa: F811
):
    """Дети наследуют ставку, а сумма их налога сходится с родителем до копейки.

    1200 с НДС 20% на три точки — это 400 у каждого и 66,67 + 66,67 + 66,66
    налога. Считай ребёнок налог от своей доли сам, вышло бы 200,01, и сумма без
    НДС у детей разошлась бы с родителем — теми самыми копейками, которые потом
    никто не объяснит.
    """
    login_as(client, "admin")
    set_rule(client, item, method="even")
    client.post("/logout/")

    login_as(client, "director")
    try:
        spend(client, item, units, unit="network", amount="1200.00", vat_rate="20")

        rows = sql.execute(
            """select f.allocation::text, f.amount, f.vat_rate, f.vat_amount
                 from facts f
                where f.expense_item_id = %s and f.superseded_at is null
                order by f.allocation, f.amount""",
            (item,),
        ).fetchall()
        children = [row for row in rows if row[0] == "allocated"]
        parent = [row for row in rows if row[0] == "split"]

        assert len(children) == 3, rows
        assert sum(row[1] for row in children) == D("1200.00"), children
        assert {row[2] for row in children} == {D("20.000")}, children
        assert sum(row[3] for row in children) == D("200.00"), children
        assert parent and parent[0][3] == D("200.00"), parent
    finally:
        client.post("/logout/")


# --- сторно: налог отменяется вместе с суммой ------------------------------------


def test_the_storno_of_an_expense_reverses_its_vat_as_well(
    client, sql, item, units, facts_removed,  # noqa: F811
):
    """Отменяющая строка несёт отрицательный налог, а не положительный.

    Иначе сторно отменяло бы сумму документа и **добавляло** налог: сумма без
    НДС за месяц уехала бы на два налога вместо нуля.
    """
    login_as(client, "director")
    try:
        key = spend(client, item, units, unit=units["NS1"], amount="1200.00", vat_rate="20")
        fact_id = sql.execute(
            "select id from facts where dedup_key = %s and superseded_at is null",
            (f"manual:cash:{key}",),
        ).fetchone()[0]

        answer = client.post(f"/expenses/{fact_id}/delete/")
        assert answer.status_code in (302, 200), body(answer)

        rows = sql.execute(
            """select amount, vat_amount from facts
                where expense_item_id = %s and superseded_at is null
                order by amount""",
            (item,),
        ).fetchall()
        assert sum(row[0] for row in rows) == D("0.00"), rows
        assert sum(row[1] or D(0) for row in rows) == D("0.00"), rows
    finally:
        client.post("/logout/")
