"""Прочитать скачанные смоуком книги обратно и сверить их с базой (T032).

Смоук нажимает ссылки браузером, а числа проверяются здесь: «файл скачался» не
означает ничего — он может быть пустым, обрезанным или собранным из другой
выборки. Открывается он тем же openpyxl, которым его откроет человек в Excel.

Сверяется файл **не сам с собой**: итог книги сравнивается с тем, что база
отдаёт этой же роли, и спрашивается она ролью `app_user`. Под владельцем схемы
политики не действуют, и сошлось бы даже при снятых.

    python tools/check_smoke_exports.py /tmp/rep3-downloads

Ждёт подпапку на роль (`director/`, `accountant/`) и в каждой три книги.
"""
from __future__ import annotations

import os
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl
import psycopg

PERIOD = "2026-06-01"

# Ориентиры приёмки на данных сида — те же, что у ведомости на экране (T028).
CONTROL = {
    "director": Decimal("1951806.13"),
    "accountant": Decimal("464752.41"),
}

# Чужой регистр не должен встретиться в файле ни строкой, ни названием (D023).
HIDDEN_FROM_ACCOUNTANT = ("Дополнительный", "Внутренний", "supplementary", "internal")

FILES = ("payout-2026-06.xlsx", "pnl-2026-06.xlsx", "partner-2026-06.xlsx")

failed: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"{'OK  ' if ok else 'FAIL'}  {name}{' — ' + detail if detail else ''}")
    if not ok:
        failed.append(name)


def text_of(book) -> str:
    return "\n".join(
        str(value)
        for ws in book.worksheets
        for row in ws.iter_rows(values_only=True)
        for value in row
        if value is not None
    )


def column(ws, header: str) -> list:
    index, out = None, []
    for row in ws.iter_rows(values_only=True):
        cells = [str(v) if v is not None else None for v in row]
        if index is not None and row[index] is not None:
            out.append(row[index])
        if header in cells:
            index = cells.index(header)
    return out


def visible_total(dsn: str, user_id: str) -> Decimal:
    """Сколько база отдаёт этой роли — её же ролью, а не владельцем схемы."""
    with psycopg.connect(dsn) as conn, conn.cursor() as cur:
        cur.execute("set local role app_user")
        cur.execute("select set_config('app.user_id', %s, true)", (user_id,))
        cur.execute(
            """select coalesce(sum(c.amount), 0)
                 from pay_components c
                 join payslips p on p.id = c.payslip_id
                 join payruns r on r.id = p.payrun_id
                where r.period = %s""",
            (PERIOD,),
        )
        return cur.fetchone()[0]


def main() -> int:
    root = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/rep3-downloads")
    dsn = os.environ["DATABASE_URL"]

    # Идентификаторы учёток сида считает сам сид — повторять его формулу здесь
    # значило бы завести вторую правду о том, кто есть кто.
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    import django

    django.setup()
    from core.management.commands.seed_dev import det_id

    for user, expected in CONTROL.items():
        folder = root / user
        for name in FILES:
            check(f"{user}: скачан {name}", (folder / name).exists(),
                  str(folder / name))
        if not all((folder / name).exists() for name in FILES):
            # Дальше читать нечего, но остальные роли проверить надо: молча
            # выйти значило бы отчитаться о неполном прогоне как о полном.
            continue

        book = openpyxl.load_workbook(folder / "payout-2026-06.xlsx")
        totals = [Decimal(str(v)) for v in column(book.active, "Итого")]
        *lines, footer = totals

        check(f"{user}: подвал файла равен сумме его строк",
              sum(lines, Decimal(0)) == footer,
              f"{sum(lines, Decimal(0))} vs {footer}")
        check(f"{user}: в файле {expected}", footer == expected, str(footer))

        from_db = visible_total(dsn, str(det_id("user", user)))
        check(f"{user}: файл равен тому, что база отдаёт его ролью",
              footer == from_db, f"файл {footer}, база {from_db}")

        pnl = openpyxl.load_workbook(folder / "pnl-2026-06.xlsx")
        kinds = column(pnl.active, "Тип строки")
        amounts = column(pnl.active, "Сумма")
        accruals = sum(
            (Decimal(str(a)) for k, a in zip(kinds, amounts, strict=True)
             if k == "Начисление"),
            Decimal(0),
        )
        check(f"{user}: начисления в P&L равны ведомости", accruals == expected,
              str(accruals))

        partner = openpyxl.load_workbook(folder / "partner-2026-06.xlsx")
        check(f"{user}: в виде бухгалтера есть листы", bool(partner.worksheets),
              ", ".join(partner.sheetnames))

        if user == "accountant":
            for name in FILES:
                text = text_of(openpyxl.load_workbook(folder / name))
                for word in HIDDEN_FROM_ACCOUNTANT:
                    check(f"accountant/{name}: слова «{word}» в файле нет",
                          word not in text)

    print("\n" + "=" * 60)
    print(f"провалено проверок: {len(failed)}")
    if failed:
        print("ПРОВАЛЕНО: " + "; ".join(failed))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
