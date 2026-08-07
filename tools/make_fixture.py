"""
Генератор обезличенной фикстуры формата PLATA.

Зачем. Регрессия должна гоняться у всех и в CI, но настоящая таблица партнёра
содержит ФИО, ставки и суммы живых людей — в репозиторий она не попадает никогда.
Поэтому здесь собирается файл той же структуры с выдуманными людьми: те же восемь
листов, те же заголовки, все четыре схемы расчёта.

Чего этот файл НЕ проверяет. Ожидаемые значения в нём посчитаны самим движком,
поэтому он ловит изменение поведения (регрессию), но не расхождение с бухгалтерией.
Сверка с реальным расчётом — отдельный тест, он берёт настоящую таблицу по пути
из переменной PAYROLL_FIXTURE и пропускается, если её нет.

Запуск:
    python tools/make_fixture.py            # перезаписать tests/fixtures/plata-sample.xlsx
"""
from __future__ import annotations

import sys
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from payroll import Employee, PayrollEngine, Timesheet, load_preset  # noqa: E402
from payroll.importers.plata_xlsx import GROSS_HEADER, SHEET_MAP  # noqa: E402

OUT = ROOT / "tests" / "fixtures" / "plata-sample.xlsx"

# Выдуманные люди. Имена сербские по звучанию, но не принадлежат никому из
# сотрудников партнёра — это принципиально, а не случайно.
FIRST = ["MARKO", "JELENA", "NIKOLA", "ANA", "STEFAN", "MILICA", "DUSAN", "TIJANA",
         "VUK", "SANJA", "LAZAR", "IVANA", "OGNJEN", "TAMARA", "PETAR", "MAJA",
         "FILIP", "KATARINA", "BOJAN", "NEVENA", "UROS", "SOFIJA", "DAMJAN", "LENA",
         "ALEKSA", "DUNJA", "VELJKO", "MINA", "STRAHINJA", "TEODORA", "PAVLE", "SARA"]
LAST = ["JOVANOVIC", "PETROVIC", "NIKOLIC", "MARKOVIC", "ILIC", "PAVLOVIC",
        "STOJANOVIC", "LUKIC", "MILOSEVIC", "KOSTIC", "TODOROVIC", "RISTIC",
        "MITROVIC", "SIMIC", "ZIVKOVIC", "VUKOVIC", "RADIC", "BOGDANOVIC",
        "SAVIC", "DIMITRIJEVIC", "ANDRIC", "OBRADOVIC", "MILADINOVIC", "VASIC",
        "KRSTIC", "TOMIC", "MAKSIMOVIC", "GAJIC", "BABIC", "LAZIC", "NENADOVIC",
        "ARSIC"]

# Колонки в порядке записи. Заголовки обязаны совпадать с теми, что ищет импортёр:
# см. FIELDS в src/payroll/importers/plata_xlsx.py.
COLUMNS = [
    ("№", None),
    ("IME", None),
    ("PREZIME", None),
    ("KOEFICIJENT", "coefficient"),
    ("CENA RADA", "base_rate"),
    ("UKUPNO SATI ZA OBRACUN DOPRINOSA", "insured"),
    ("SATI RADA ZA OBRACUN NETA", "regular"),
    ("SAT NA RAD PRAZNIKA", "holiday"),
    ("GODISNJI ODMOR", "vacation"),
    ("SATI BOLOVANJE", "sick"),
    ("KOREKCIJA DO MINIMALCA", "correction"),
    ("OBUSTAVA", "deduction"),
    ("ISPLATA U KES", "cash"),
    ("TOPLI OBROK I REGRES", "meal"),
    ("UKUPNO ZA ISPLATU", "exp_net"),
    ("__BRUTO__", "exp_gross"),          # заголовок подставляется по схеме
    ("DOPRINOSI", "exp_contrib"),
    ("UKUPAN TROSAK PO RADNIKU", "exp_total"),
]

# Часы по схемам: (обычные, праздничные, отпуск, больничный).
# Варианты подобраны так, чтобы задеть разные ветки движка: чистый месяц,
# праздник, отпуск, больничный с доплатой до минимума.
HOURS = {
    "standard": [(176, 0, 0, 0), (152, 8, 16, 0), (148, 0, 8, 20), (160, 8, 8, 0)],
    "half_time": [(88, 0, 0, 0), (80, 0, 8, 0), (60, 8, 0, 20), (20, 0, 0, 0)],
    "half_time_min_base": [(88, 0, 0, 0), (72, 8, 8, 0), (88, 0, 0, 0), (64, 0, 16, 8)],
    "temporary": [(120, 0, 0, 0), (96, 0, 0, 0), (64, 0, 0, 0), (40, 0, 0, 0)],
}

COEFFICIENTS = [Decimal("1.0"), Decimal("1.135"), Decimal("1.181"), Decimal("1.25")]
BASE_RATE = Decimal("371")


def money(value: Decimal) -> float:
    return float(Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def build() -> None:
    engine = PayrollEngine(load_preset("serbia-2026"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    person = 0
    for sheet_name, (scheme, group) in SHEET_MAP.items():
        ws = wb.create_sheet(sheet_name[:31])

        headers = [
            GROSS_HEADER[scheme][0] if title == "__BRUTO__" else title
            for title, _ in COLUMNS
        ]
        ws.append(headers)

        for n, (regular, holiday, vacation, sick) in enumerate(HOURS[scheme], start=1):
            coefficient = COEFFICIENTS[person % len(COEFFICIENTS)]
            first, last = FIRST[person % len(FIRST)], LAST[person % len(LAST)]
            person += 1

            insured = regular + holiday + vacation + sick
            employee = Employee(
                ext_id=f"{first} {last}", name=f"{first} {last}",
                group=group, scheme=scheme,
                base_rate=BASE_RATE, coefficient=coefficient,
            )
            timesheet = Timesheet(
                hours={"regular": Decimal(regular), "holiday": Decimal(holiday),
                       "vacation": Decimal(vacation), "sick": Decimal(sick)},
                insured_hours=Decimal(insured),
                norm_hours=Decimal(insured),
            )

            slip = engine.calculate(employee, timesheet)

            values = {
                "coefficient": float(coefficient),
                "base_rate": float(BASE_RATE),
                "insured": insured,
                "regular": regular,
                "holiday": holiday,
                "vacation": vacation,
                "sick": sick,
                "correction": None,
                "deduction": None,
                "cash": None,
                # Колонка надбавки в обезличенном файле остаётся пустой намеренно.
                # В настоящей таблице бухгалтер иногда проставляет её руками, и
                # проверка подменяет расчётное значение табличным — чтобы сверять
                # схемы расчёта, а не ввод. Но если заполнить её здесь, подмена
                # сработает на всех строках, и регрессия перестанет замечать
                # изменение самого правила начисления: правку ставки надбавки
                # тест пропускал молча. Проверено порчей 2026-08-07.
                "meal": None,
                "exp_net": money(slip.net),
                "exp_gross": money(slip.gross),
                "exp_contrib": money(slip.contributions),
                "exp_total": money(slip.total_cost),
            }
            ws.append([n, first, last] + [values[key] for _, key in COLUMNS[3:]])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"записано: {OUT.relative_to(ROOT)}, листов {len(wb.sheetnames)}, строк {person}")


if __name__ == "__main__":
    build()
